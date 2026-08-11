"""Tests for Exposure condition-table population + Frictionless bridge (Issue #144).

``populate_condition_table`` writes per-well rows into the Exposure condition
table CSV (replacing the header-only placeholder from #94). ``csvw_to_frictionless``
converts the CSVW column descriptors into the Frictionless ``{fields: [...]}``
shape so ``validate_table`` needs no hand-authored schema.
"""

from __future__ import annotations

import csv
from pathlib import Path

import pytest

from builder.state import CrateState, Entity, EntityProvenance
from builder.tools._crate_mapping import _CONDITION_TABLE_COLUMNS
from builder.tools.data_content import (
    _UNIT_SUFFIX_RE,
    csvw_to_frictionless,
    populate_condition_table,
    project_condition_rows,
    validate_table,
)

# Every test here exports a crate, and each export now runs the uncached,
# owlrl-heavy validator over all three profiles at the full severity gate (#446)
# — ~10s per export locally, and the 2-vCPU CI runner is ~2-3x slower, which puts
# the whole module against the CI-wide `--timeout=30`. Same headroom, for the
# same reason, that the other export-heavy modules already take
# (test_export_smoke, test_readers, test_path_traversal, test_html_xss).
# Headroom, not a licence to grow: no test in this module is changed.
pytestmark = pytest.mark.timeout(120)

# The committed per-well fixture the pipeline actually meets in the corpus. Anchored
# to this file, not the CWD — a test below deliberately chdirs. The fixture is owned
# by the eval corpus, so ``test_the_corpus_fixture_still_requires_aliasing`` guards
# against it drifting to canonical column names and quietly making the end-to-end
# case below vacuous.
_FIXTURE_CSV = (
    Path(__file__).parent / "fixtures" / "svhps22_input" / "raw_data" / "dose_response_raw.csv"
)

# The real depositor file behind #471: the corpus's only genuine tidy per-well
# design table, 1048 rows, fifteen headers, none of them canonical.
_TIDY_FIXTURE_CSV = (
    Path(__file__).parent
    / "fixtures"
    / "svhps22_real_input"
    / "assay_01_TH_uptake"
    / "EDCs"
    / "Combined uptake data EDCs_tidy.csv"
)


def _tidy_rows(limit: int | None = None) -> list[dict[str, str]]:
    """Rows of the real tidy fixture, read off disk — never hand-authored here."""
    with open(_TIDY_FIXTURE_CSV, newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    return rows[:limit] if limit is not None else rows


def _exposure_state() -> CrateState:
    state = CrateState()
    state.metadata.title = "Exposure crate"
    state.add_entity(
        Entity(
            entity_id="proc_exp",
            type="LabProcess",
            fields={"process_type": "Exposure", "name": "Exposure step"},
            _provenance=EntityProvenance(created_by="llm"),
        )
    )
    return state


def test_csvw_to_frictionless_maps_columns():
    schema = csvw_to_frictionless(_CONDITION_TABLE_COLUMNS)
    assert "fields" in schema
    names = [f["name"] for f in schema["fields"]]
    # The full 10-column condition-table schema (Issue #180, Lane D).
    assert names == [
        "well_id",
        "assay",
        "cell_line",
        "compound",
        "concentration_value",
        "concentration_unit",
        "exposure_duration",
        "experiment",
        "technical_replicate",
        "control",
    ]
    by_name = {f["name"]: f for f in schema["fields"]}
    # double -> number, string -> string (Frictionless types)
    assert by_name["concentration_value"]["type"] == "number"
    assert by_name["cell_line"]["type"] == "string"


def test_populate_condition_table_writes_rows(tmp_path):
    state = _exposure_state()
    # Population fills whatever columns the data provides; the schema describes
    # all 10, missing columns are written empty (extrasaction="ignore").
    rows = [
        {"well_id": "A1", "cell_line": "HepG2", "compound": "Aspirin",
         "concentration_value": "10", "concentration_unit": "uM",
         "exposure_duration": "24h"},
        {"well_id": "A2", "cell_line": "HepG2", "compound": "Aspirin",
         "concentration_value": "100", "concentration_unit": "uM",
         "exposure_duration": "24h"},
    ]
    result = populate_condition_table(state, "proc_exp", rows, output_dir=str(tmp_path))
    assert result["ok"] is True
    csv_path = result["path"]
    with open(csv_path, newline="", encoding="utf-8") as fh:
        reader = list(csv.DictReader(fh))
    assert len(reader) == 2
    assert reader[0]["concentration_value"] == "10"
    assert reader[1]["concentration_value"] == "100"
    assert reader[0]["well_id"] == "A1"


def test_populated_table_validates_with_inferred_schema(tmp_path):
    state = _exposure_state()
    rows = [
        {"well_id": "A1", "cell_line": "HepG2", "compound": "Aspirin",
         "concentration_value": "10", "concentration_unit": "uM",
         "exposure_duration": "24h"},
    ]
    result = populate_condition_table(state, "proc_exp", rows, output_dir=str(tmp_path))
    schema = csvw_to_frictionless(_CONDITION_TABLE_COLUMNS)
    report = validate_table(result["path"], schema)
    assert report["ok"] is True, report


def test_populated_table_rejects_non_numeric_concentration(tmp_path):
    """The inferred CSVW schema TYPES ``concentration_value`` as a number, so a row
    whose value is valid-as-string but not-a-number must FAIL content validation with
    an error routed to that column.

    The happy-path test above would still pass even if the column were typed as a
    plain string — this negative case is what makes the number-typing meaningful.
    """
    state = _exposure_state()
    rows = [
        {"well_id": "A1", "cell_line": "HepG2", "compound": "Aspirin",
         "concentration_value": "abc",  # not a number — must be rejected by the type
         "concentration_unit": "uM", "exposure_duration": "24h"},
    ]
    result = populate_condition_table(state, "proc_exp", rows, output_dir=str(tmp_path))
    schema = csvw_to_frictionless(_CONDITION_TABLE_COLUMNS)
    report = validate_table(result["path"], schema)

    assert report["ok"] is False, report
    assert any(
        "concentration_value" in str(issue.get("property", ""))
        or "concentration_value" in str(issue.get("message", ""))
        for issue in report["issues"]
    ), report["issues"]


class TestProjectConditionRows:
    """Source headers are aliased onto the ten canonical titles (#381a).

    ``populate_condition_table`` writes through ``csv.DictWriter(...,
    extrasaction="ignore")``, so any source column not named *exactly* like a
    canonical title was silently dropped and its canonical slot written empty.
    Real plate maps do not use the canonical names.
    """

    def test_projects_the_committed_fixture_header(self) -> None:
        # The real corpus fixture: well,compound,cell_line,concentration_uM,
        # tpo_activity_rfu. Only 2 of 10 titles matched before aliasing, so the
        # dose axis — the entire point of the table — landed blank.
        rows = [
            {
                "well": "A1",
                "compound": "Methimazole",
                "cell_line": "FRTL-5",
                "concentration_uM": "0.3",
                "tpo_activity_rfu": "15110",
            }
        ]
        result = project_condition_rows(rows)
        got = result["rows"][0]
        assert got["well_id"] == "A1"
        assert got["compound"] == "Methimazole"
        assert got["cell_line"] == "FRTL-5"
        assert got["concentration_value"] == "0.3"
        assert got["concentration_unit"] == "uM"

    def test_reports_measurement_columns_as_unmapped_never_swallowed(self) -> None:
        rows = [{"well": "A1", "tpo_activity_rfu": "15110"}]
        result = project_condition_rows(rows)
        assert "tpo_activity_rfu" in result["unmapped_source_columns"]

    def test_suffix_unit_rule_splits_value_and_unit(self) -> None:
        result = project_condition_rows([{"well": "A1", "dose_mM": "2.5"}])
        assert result["rows"][0]["concentration_value"] == "2.5"
        assert result["rows"][0]["concentration_unit"] == "mM"

    def test_unit_stays_a_literal_string_never_an_ontology_iri(self) -> None:
        # D5: normalising a unit to a UO IRI needs an authoritative lookup. The
        # suffix is prose from a filename — it is carried verbatim, never lifted.
        unit = project_condition_rows([{"well": "A1", "conc_uM": "1"}])["rows"][0][
            "concentration_unit"
        ]
        assert unit == "uM"
        assert "://" not in unit and ":" not in unit

    def test_canonical_name_wins_over_an_alias(self) -> None:
        # Both present: the canonical column is authoritative, the alias is ignored.
        rows = [{"well_id": "CANON", "well": "ALIAS"}]
        assert project_condition_rows(rows)["rows"][0]["well_id"] == "CANON"

    def test_alias_does_not_overwrite_a_populated_canonical_value(self) -> None:
        rows = [{"concentration_value": "10", "concentration_unit": "nM", "dose_uM": "99"}]
        got = project_condition_rows(rows)["rows"][0]
        assert got["concentration_value"] == "10"
        assert got["concentration_unit"] == "nM"

    def test_reports_which_canonical_columns_were_mapped(self) -> None:
        result = project_condition_rows([{"well": "A1", "chemical": "Aspirin"}])
        assert set(result["mapped_columns"]) == {"well_id", "compound"}

    def test_the_stale_react_tool_column_names_now_land(self) -> None:
        # react/tools_spec.py advertised the pre-#180 names concentration / unit /
        # duration. All three were dropped by extrasaction="ignore", so a model
        # obeying the tool description verbatim wrote an empty table.
        rows = [
            {
                "well": "A1",
                "cell_line": "HepG2",
                "compound": "Aspirin",
                "concentration": "10",
                "unit": "uM",
                "duration": "24h",
            }
        ]
        got = project_condition_rows(rows)["rows"][0]
        assert got["concentration_value"] == "10"
        assert got["concentration_unit"] == "uM"
        assert got["exposure_duration"] == "24h"


class TestPopulateRefusesRatherThanBlanking:
    """Writing a table of blank cells is worse than the honest header (#381a)."""

    def _dest(self, tmp_path: Path) -> Path:
        return tmp_path / "data" / "proc_exp_condition_table.csv"

    def test_refuses_when_no_canonical_column_maps(self, tmp_path) -> None:
        state = _exposure_state()
        rows = [{"instrument": "plate reader", "operator": "JH"}]
        result = populate_condition_table(state, "proc_exp", rows, output_dir=str(tmp_path))
        assert result["ok"] is False
        assert result["unmapped_source_columns"]

    def test_refusal_writes_nothing_at_all(self, tmp_path) -> None:
        # The critical half: a refusal must not leave a blanked file behind, and
        # must not clobber a header the build already wrote.
        state = _exposure_state()
        populate_condition_table(
            state, "proc_exp", [{"instrument": "plate reader"}], output_dir=str(tmp_path)
        )
        assert not self._dest(tmp_path).exists()

    def test_refuses_when_no_row_resolves_a_well_id(self, tmp_path) -> None:
        # A per-well design table without a well key is not a design table.
        state = _exposure_state()
        rows = [{"compound": "Aspirin", "cell_line": "HepG2"}]
        result = populate_condition_table(state, "proc_exp", rows, output_dir=str(tmp_path))
        assert result["ok"] is False
        assert not self._dest(tmp_path).exists()

    def test_success_still_reports_the_skipped_columns(self, tmp_path) -> None:
        state = _exposure_state()
        rows = [{"well": "A1", "compound": "Aspirin", "tpo_activity_rfu": "18420"}]
        result = populate_condition_table(state, "proc_exp", rows, output_dir=str(tmp_path))
        assert result["ok"] is True
        assert "tpo_activity_rfu" in result["unmapped_source_columns"]

    def test_the_corpus_fixture_still_requires_aliasing(self) -> None:
        """Anti-drift guard for the end-to-end case below.

        The fixture belongs to the eval corpus, not to this test. If its header
        ever became canonical, the end-to-end assertion would still pass while
        exercising no aliasing at all — a tautology. Pin the two properties the
        end-to-end case actually depends on: the header is non-canonical, and it
        carries a suffix-unit dose column.
        """
        with open(_FIXTURE_CSV, newline="", encoding="utf-8") as fh:
            header = next(csv.reader(fh))
        canonical = {c["titles"] for c in _CONDITION_TABLE_COLUMNS}
        assert set(header) - canonical, (
            f"fixture header {header} is now fully canonical — the end-to-end test "
            "below no longer exercises aliasing and must be rewritten"
        )
        assert "well" in header and "well_id" not in header
        assert any(c.startswith(("concentration_", "conc_", "dose_")) for c in header)

    def test_aliased_plate_map_lands_real_values_end_to_end(self, tmp_path) -> None:
        # The whole point: the committed fixture, read from disk, produces a table
        # whose dose axis is populated rather than blank.
        state = _exposure_state()
        result = populate_condition_table(
            state, "proc_exp", str(_FIXTURE_CSV), output_dir=str(tmp_path)
        )
        assert result["ok"] is True, result
        with open(result["path"], newline="", encoding="utf-8") as fh:
            written = list(csv.DictReader(fh))
        assert len(written) > 1
        assert [r["concentration_value"] for r in written] != [""] * len(written)
        assert all(r["well_id"] for r in written)
        assert {r["concentration_unit"] for r in written} == {"uM"}


class TestTidyExportVocabulary:
    """The S-VHPS22 tidy per-well export is aliased onto the canonical ten (#471).

    ``Combined uptake data EDCs_tidy.csv`` is the only genuine tidy per-well design
    table in the nine-dataset corpus, and not one of its fifteen headers matched a
    canonical title, an alias or the suffix-unit rule — so ``mapped_columns`` came
    back empty, the first refusal gate fired, and the study shipped the header-only
    placeholder while 1048 real rows sat on disk.

    These drive the real projection over the real file, never a paraphrase of it.
    """

    def test_the_tidy_header_can_only_be_mapped_by_an_alias(self) -> None:
        """Honesty control for everything below.

        If a single tidy header were already canonical (or matched the suffix-unit
        rule) the assertions below would pass while proving nothing about the alias
        table. Pin that neither older path can reach this file: whatever lands in
        the projection got there through the #471 alias entries.
        """
        with open(_TIDY_FIXTURE_CSV, newline="", encoding="utf-8") as fh:
            header = next(csv.reader(fh))
        canonical = {c["titles"] for c in _CONDITION_TABLE_COLUMNS}
        assert not set(header) & canonical, (
            f"tidy header {header} now shares a name with a canonical column — the "
            "alias assertions below no longer prove the alias table did the work"
        )
        assert not [h for h in header if _UNIT_SUFFIX_RE.match(h)], (
            "a suffix-unit header would populate the dose axis without any alias"
        )

    def test_the_real_header_populates_the_design_columns(self) -> None:
        rows = _tidy_rows(limit=8)
        projected = project_condition_rows(rows)["rows"][0]
        source = rows[0]
        # Right-hand sides are read out of the file, so each assertion says "this
        # canonical cell carries the value of that source column" — the mapping
        # claim itself — rather than restating a literal the test chose.
        assert projected["compound"] == source["test_substance_id"]
        assert projected["cell_line"] == source["biosample_type"]
        assert projected["concentration_value"] == source["exposure_concentration_value"]
        assert projected["concentration_unit"] == source["exposure_concentration_unit"]
        assert projected["assay"] == source["assay_endpoint"]
        assert projected["technical_replicate"] == source["replicate_id"]
        # A tidy export has no plate geometry; run_id is its row key and well_id is
        # typed dcterms:identifier, so the source's own key goes there unaltered.
        # Without it no row resolves a well_id and the SECOND refusal gate fires,
        # which is why the issue's "one alias per column" sketch was not enough.
        assert projected["well_id"] == source["run_id"]

    def test_measurement_columns_are_still_reported_never_swallowed(self) -> None:
        unmapped = set(project_condition_rows(_tidy_rows(limit=8))["unmapped_source_columns"])
        # Results and prose are not design. Aliasing them into a design column
        # would be fabrication (D5); reporting them is the honest outcome.
        assert {
            "measurement_type",
            "measurement_value",
            "measurement_unit",
            "measurement_date",
            "notes",
        } <= unmapped
        # biosample_id is left unaliased on purpose: it would collide with
        # biosample_type on cell_line, and the file never fills it.
        assert "biosample_id" in unmapped

    def test_populate_writes_the_table_it_used_to_refuse(self, tmp_path) -> None:
        state = _exposure_state()
        result = populate_condition_table(
            state, "proc_exp", str(_TIDY_FIXTURE_CSV), output_dir=str(tmp_path)
        )
        assert result["ok"] is True, result
        with open(result["path"], newline="", encoding="utf-8") as fh:
            written = list(csv.DictReader(fh))
        source = _tidy_rows()
        assert len(written) == len(source)
        assert all(r["well_id"] for r in written)
        assert {r["compound"] for r in written} == {r["test_substance_id"] for r in source}
        assert {r["cell_line"] for r in written} == {r["biosample_type"] for r in source}
        assert {r["concentration_unit"] for r in written} == {
            r["exposure_concentration_unit"] for r in source
        }
        # Success must not quietly absorb the measurement columns it skipped.
        assert "measurement_value" in result["unmapped_source_columns"]

    def test_a_split_duration_is_composed_rather_than_half_dropped(self) -> None:
        # Both halves mean `exposure_duration`, which is ONE string column. Aliasing
        # both would make them collide and keep only "24" — and h vs d is a 24x
        # error, the same magnitude trap the uM/mM suffix rule exists to avoid.
        rows = [{"well": "A1", "exposure_duration_value": "24", "exposure_duration_unit": "h"}]
        duration = project_condition_rows(rows)["rows"][0]["exposure_duration"]
        assert duration == "24 h"

    def test_a_composed_duration_stays_a_literal_never_an_ontology_iri(self) -> None:
        rows = [{"exposure_duration_value": "24", "exposure_duration_unit": "h"}]
        duration = project_condition_rows(rows)["rows"][0]["exposure_duration"]
        assert "://" not in duration and ":" not in duration

    def test_a_canonical_duration_outranks_the_split_pair(self) -> None:
        rows = [
            {
                "exposure_duration": "48h",
                "exposure_duration_value": "24",
                "exposure_duration_unit": "h",
            }
        ]
        assert project_condition_rows(rows)["rows"][0]["exposure_duration"] == "48h"

    def test_a_duration_unit_with_no_magnitude_writes_nothing_and_says_so(self) -> None:
        # "h" on its own describes no duration, so no cell is written. The header
        # IS reported, though: the pair rule understood it and still discarded a
        # real value, and `unmapped_source_columns` is the only channel that can
        # tell the caller so. Suppressing it here — on the grounds that the header
        # was "understood" — is how a two-day exposure would ship as a bare `2`
        # with the `d` gone and nothing left to notice.
        result = project_condition_rows([{"well": "A1", "exposure_duration_unit": "h"}])
        assert "exposure_duration" not in result["rows"][0]
        assert "exposure_duration_unit" in result["unmapped_source_columns"]

    def test_a_unit_orphaned_by_another_alias_is_still_reported(self) -> None:
        # The case that makes the rule above load-bearing rather than pedantic:
        # `exposure_time` fills `exposure_duration` through the ALIAS pass, so the
        # pair never composes and the `d` is dropped. A 2-day exposure must not
        # ship as "2" in silence.
        result = project_condition_rows(
            [{"well": "A1", "exposure_time": "2", "exposure_duration_unit": "d"}]
        )
        assert result["rows"][0]["exposure_duration"] == "2"
        assert "exposure_duration_unit" in result["unmapped_source_columns"]

    def test_a_blank_duration_pair_is_not_reported(self) -> None:
        # Nothing was discarded, so there is nothing to report: an empty row is
        # what the alias pass does with a blank value too. This is the control
        # that stops the rule above from degenerating into "always report".
        result = project_condition_rows(
            [{"well": "A1", "exposure_duration_value": "", "exposure_duration_unit": ""}]
        )
        assert "exposure_duration" not in result["rows"][0]
        assert result["unmapped_source_columns"] == []

    def test_two_aliases_for_one_column_resolve_in_source_order(self) -> None:
        # `chemical` and `test_substance_id` both mean `compound`. A row carrying
        # both is a real hazard, so the outcome is defined rather than incidental:
        # the first source column wins, in the source's own column order.
        both = {"chemical": "Aspirin", "test_substance_id": "Methimazole"}
        assert project_condition_rows([both])["rows"][0]["compound"] == "Aspirin"
        swapped = dict(reversed(list(both.items())))
        result = project_condition_rows([swapped])
        assert result["rows"][0]["compound"] == "Methimazole"
        # The loser had a canonical home; only homeless columns belong in that list.
        assert result["unmapped_source_columns"] == []


class TestPopulatePathResolution:
    def test_falls_back_to_the_session_crate_path_not_cwd(self, tmp_path, monkeypatch) -> None:
        # data_content resolved a missing output_dir to Path.cwd() while
        # export_crate resolves it to _default_crate_path(state). Rows written to
        # the wrong root are silently lost at export.
        from builder.tools.builder import _default_crate_path

        state = _exposure_state()
        state.session_id = "sess-381"
        monkeypatch.chdir(tmp_path)
        result = populate_condition_table(
            state, "proc_exp", [{"well": "A1", "compound": "Aspirin"}]
        )
        assert result["ok"] is True, result
        assert Path(result["path"]).is_relative_to(Path(_default_crate_path(state)))


class TestReactToolDescriptionMatchesTheCode:
    """The advertised columns must be the real ones (#381a).

    The ReAct description named the pre-#180 five-column set long after those
    columns stopped existing, so a model that obeyed it verbatim had every value
    discarded by ``extrasaction="ignore"``. Nothing tested the description, which
    is precisely how it drifted. These pin it to the code.
    """

    def _spec(self) -> dict:
        from builder.agents.react.tools_spec import TOOL_SPECS

        return next(t for t in TOOL_SPECS if t["name"] == "populate_condition_table")

    def test_every_canonical_column_is_advertised(self) -> None:
        text = str(self._spec())
        for column in (c["titles"] for c in _CONDITION_TABLE_COLUMNS):
            assert column in text, f"tool description does not mention {column!r}"

    def test_retired_column_names_are_not_advertised_as_keys(self) -> None:
        # "concentration"/"unit"/"duration" survive only as substrings of the
        # canonical names, so assert on the delimited forms a model would copy.
        text = str(self._spec())
        for retired in ("concentration/unit/duration", "concentration/unit", "unit/duration"):
            assert retired not in text, f"stale column list {retired!r} still advertised"

    def test_the_tidy_export_vocabulary_is_advertised(self) -> None:
        # Same drift trap as the retired five-column names: an alias the model is
        # never told about is an alias it will not aim for when it hand-builds rows.
        from builder.tools.data_content import _CONDITION_TABLE_ALIASES

        with open(_TIDY_FIXTURE_CSV, newline="", encoding="utf-8") as fh:
            header = next(csv.reader(fh))
        understood = [h for h in header if h.lower() in _CONDITION_TABLE_ALIASES]
        assert understood, "no tidy header is aliased at all — #471 has regressed"
        text = str(self._spec())
        for name in understood:
            assert name in text, f"tool description does not mention the {name!r} alias"

    def test_the_refusal_contract_is_documented(self) -> None:
        # A tool that can refuse must say so, or the caller reads a failure as a bug.
        description = self._spec()["description"].lower()
        assert "refus" in description
        assert "unmapped_source_columns" in description


@pytest.mark.parametrize(
    "source,canonical",
    [
        ("well", "well_id"),
        ("well_position", "well_id"),
        ("conc", "concentration_value"),
        ("units", "concentration_unit"),
        ("exposure_time", "exposure_duration"),
        ("cell", "cell_line"),
        ("substance", "compound"),
        ("test_item", "compound"),
        ("replicate", "technical_replicate"),
        # The tidy per-well export vocabulary (#471), header names taken from
        # S-VHPS22's Combined uptake data EDCs_tidy.csv.
        ("run_id", "well_id"),
        ("biosample_type", "cell_line"),
        ("test_substance_id", "compound"),
        ("exposure_concentration_value", "concentration_value"),
        ("exposure_concentration_unit", "concentration_unit"),
        ("assay_endpoint", "assay"),
        ("replicate_id", "technical_replicate"),
    ],
)
def test_alias_table_covers_the_documented_synonyms(source: str, canonical: str) -> None:
    got = project_condition_rows([{source: "X"}])["rows"][0]
    assert got[canonical] == "X"


class TestMultivaluedColumns:
    """#408 (c) — a populated column may no longer carry a column-wide ``valueUrl``.

    ``_build_condition_table_schema`` asserts ``cells[0]`` / ``chems[0]`` for the
    WHOLE cell_line / compound column. At zero rows that claim is vacuous. Once
    rows exist it says every value in the column resolves to that one entity — so
    populating a multi-compound plate converts unverified prose into false
    entity-resolved claims. D5 says refuse the claim, not inherit it.
    """

    def _write(self, tmp_path: Path, rows: list[dict[str, str]]) -> Path:
        csv_path = tmp_path / "condition_table.csv"
        titles = [c["titles"] for c in _CONDITION_TABLE_COLUMNS]
        with csv_path.open("w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=titles, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)
        return csv_path

    def test_a_single_valued_column_is_not_multivalued(self, tmp_path: Path) -> None:
        from builder.tools.data_content import condition_table_multivalued_columns

        csv_path = self._write(
            tmp_path,
            [
                {"well_id": "A1", "cell_line": "CHO-K1", "compound": "T4"},
                {"well_id": "A2", "cell_line": "CHO-K1", "compound": "T4"},
            ],
        )
        got = condition_table_multivalued_columns(str(csv_path))
        assert "compound" not in got
        assert "cell_line" not in got
        # Every well differs by design, so well_id IS multivalued — the function
        # reports any canonical column, and the caller decides which ones matter.
        assert got == {"well_id"}

    def test_a_multi_compound_plate_reports_the_compound_column(self, tmp_path: Path) -> None:
        from builder.tools.data_content import condition_table_multivalued_columns

        csv_path = self._write(
            tmp_path,
            [
                {"well_id": "A1", "cell_line": "CHO-K1", "compound": "T4"},
                {"well_id": "A2", "cell_line": "CHO-K1", "compound": "T3"},
            ],
        )
        got = condition_table_multivalued_columns(str(csv_path))
        assert "compound" in got
        assert "cell_line" not in got, "one cell line across both wells — still single-valued"

    def test_non_canonical_source_columns_are_ignored(self, tmp_path: Path) -> None:
        """A verbatim-copied plate map's extra headers have no CSVW column to guard."""
        from builder.tools.data_content import condition_table_multivalued_columns

        csv_path = tmp_path / "verbatim.csv"
        csv_path.write_text(
            "well_id,compound,tpo_activity_rfu\nA1,T4,101\nA2,T4,202\n", encoding="utf-8"
        )
        assert "tpo_activity_rfu" not in condition_table_multivalued_columns(str(csv_path))

    def test_blanks_do_not_count_as_a_distinct_value(self, tmp_path: Path) -> None:
        """An empty cell is absence, not a second compound."""
        from builder.tools.data_content import condition_table_multivalued_columns

        csv_path = self._write(
            tmp_path,
            [
                {"well_id": "A1", "compound": "T4"},
                {"well_id": "A2", "compound": ""},
                {"well_id": "A3", "compound": "  "},
            ],
        )
        assert "compound" not in condition_table_multivalued_columns(str(csv_path))

    def test_a_header_only_table_is_not_multivalued(self, tmp_path: Path) -> None:
        """The pre-#408 state: no rows, so nothing to contradict."""
        from builder.tools.data_content import condition_table_multivalued_columns

        assert condition_table_multivalued_columns(str(self._write(tmp_path, []))) == set()

    def test_a_missing_file_is_not_multivalued(self, tmp_path: Path) -> None:
        """The in-memory validate path has no CSV on disk — must not raise."""
        from builder.tools.data_content import condition_table_multivalued_columns

        assert condition_table_multivalued_columns(str(tmp_path / "nope.csv")) == set()


class TestValueUrlDropsOnMultivaluedColumn:
    """#408 (c), through the real crate mapping + export — not the helper alone.

    The Exposure's ``chemicals`` ref field and its cultured-sample ``object`` are
    what feed ``cells``/``chems`` into ``_build_condition_table_schema``, so these
    build the same wiring ``_materialize_plan`` produces.
    """

    _REL = Path("data") / "LabProcess_proc_exp_condition_table.csv"

    def _state(self, compound_ids: list[str]) -> CrateState:
        state = CrateState()
        state.metadata.title = "Exposure crate"

        def add(eid: str, type_: str, **fields: object) -> None:
            state.add_entity(
                Entity(
                    entity_id=eid,
                    type=type_,  # ty: ignore[invalid-argument-type]
                    fields=fields,
                    _provenance=EntityProvenance(created_by="llm"),
                )
            )

        add("samp_c", "Sample", name="Cultured CHO-K1")
        for eid in compound_ids:
            add(eid, "MolecularEntity", name=eid)
        add(
            "proc_exp",
            "LabProcess",
            process_type="Exposure",
            name="Exposure step",
            chemicals=compound_ids,
            object=["samp_c"],
        )
        return state

    def _export_value_urls(
        self, out_dir: Path, csv_body: str, compound_ids: list[str]
    ) -> dict[str, str | None]:
        """Pre-populate the condition CSV, export, and read back column valueUrls."""
        import json

        from builder.tools.builder import export_crate

        dest = out_dir / self._REL
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_text(csv_body, encoding="utf-8")

        state = self._state(compound_ids)
        state.metadata.output_path = str(out_dir)
        export_crate(state)

        doc = json.loads((out_dir / "ro-crate-metadata.json").read_text(encoding="utf-8"))
        urls: dict[str, str | None] = {}
        for node in doc["@graph"]:
            if "csvw:Column" not in str(node.get("@type", "")):
                continue
            raw = node.get("valueUrl")
            urls[str(node.get("titles") or "")] = (
                raw.get("@id") if isinstance(raw, dict) else raw
            )
        return urls

    _HEADER = "well_id,assay,cell_line,compound\n"

    def test_single_valued_compound_column_keeps_its_valueurl(self, tmp_path: Path) -> None:
        """The control: existing #94/#180 behaviour must survive the change."""
        urls = self._export_value_urls(
            tmp_path / "crate",
            self._HEADER + "A1,uptake,CHO-K1,Thyroxine\n",
            ["chem_t4"],
        )
        assert urls.get("compound"), "a single-compound plate must keep its column valueUrl"
        assert urls.get("cell_line"), "a single-cell-line plate must keep its column valueUrl"

    def test_multi_compound_column_drops_its_valueurl(self, tmp_path: Path) -> None:
        urls = self._export_value_urls(
            tmp_path / "crate",
            self._HEADER + "A1,uptake,CHO-K1,Thyroxine\nA2,uptake,CHO-K1,Triiodothyronine\n",
            ["chem_t4", "chem_t3"],
        )
        assert urls.get("compound") is None, (
            "two compounds in the column — a column-wide valueUrl would be a false claim"
        )
        # cell_line is still single-valued, so that claim stands: the guard is
        # per-column, not a blanket retreat.
        assert urls.get("cell_line"), "single-valued cell_line must keep its valueUrl"

    def test_header_only_table_keeps_its_valueurl(self, tmp_path: Path) -> None:
        """The pre-#408 vacuous case is unchanged — no rows contradict the claim."""
        urls = self._export_value_urls(tmp_path / "crate", self._HEADER, ["chem_t4"])
        assert urls.get("compound"), "a header-only table asserts nothing to contradict"

    def test_populated_rows_survive_the_export(self, tmp_path: Path) -> None:
        """`if not dest.exists()` + ro-crate-py's samefile guard must not clobber rows."""
        out = tmp_path / "crate"
        body = self._HEADER + "A1,uptake,CHO-K1,Thyroxine\nA2,uptake,CHO-K1,Triiodothyronine\n"
        self._export_value_urls(out, body, ["chem_t4", "chem_t3"])
        assert (out / self._REL).read_text(encoding="utf-8") == body


class TestPlateMapIntakeByFormat:
    """``populate_condition_table`` dispatches on the file's format (#422).

    The spine classifies a plan file as ``condition_table`` by ROLE, not by
    extension, so the real deposit's ``.xlsx`` plate map reached a UTF-8
    ``csv.DictReader`` and raised ``UnicodeDecodeError`` on the first ZIP byte.
    The spine swallowed that into a ``reason:`` string and the crate shipped a
    header-only table with nothing said about why.
    """

    def _exposure(self, tmp_path):
        from builder.state import CrateState
        from builder.tools.composites import draft_process_chain, scaffold_isa_backbone

        state = CrateState()
        state.metadata.title = "Plate map intake"
        state.metadata.output_path = str(tmp_path / "crate")
        scaffold = scaffold_isa_backbone(
            state, investigation={"name": "I"}, study={"name": "S"}, assay={"name": "A"}
        )
        draft_process_chain(state, scaffold["assay_id"], chain=[{"process_type": "Exposure"}])
        exposure = next(
            p
            for p in state.list_entities("LabProcess")
            if p.fields.get("process_type") == "Exposure"
        )
        return state, exposure.entity_id

    def _workbook(self, path) -> None:
        import openpyxl

        book = openpyxl.Workbook()
        cover = book.active
        cover.title = "Cover"
        cover.append(["Depositor", "Notes"])
        cover.append(["Lab X", "read me first"])
        plate = book.create_sheet("Plate map")
        plate.append(["well_id", "compound", "concentration_value", "concentration_unit"])
        for index, compound in enumerate(["BPA", "Lesinurad", "Quercetin"], start=1):
            plate.append([index, compound, 10.0, "uM"])
        book.save(path)

    def test_xlsx_plate_map_is_read(self, tmp_path) -> None:
        from builder.tools.data_content import populate_condition_table

        state, exposure_id = self._exposure(tmp_path)
        book = tmp_path / "platemap.xlsx"
        self._workbook(book)
        out = populate_condition_table(state, exposure_id, str(book))
        assert out["ok"], out
        assert out["rows"] == 3

    def test_the_data_sheet_is_chosen_not_the_first(self, tmp_path) -> None:
        # A depositor workbook opens on a cover page; "first sheet" would read it.
        from builder.tools.data_content import populate_condition_table

        state, exposure_id = self._exposure(tmp_path)
        book = tmp_path / "platemap.xlsx"
        self._workbook(book)
        assert populate_condition_table(state, exposure_id, str(book))["sheet"] == "Plate map"

    def test_integral_floats_do_not_become_1_point_0(self, tmp_path) -> None:
        # openpyxl types every numeric cell as float; a well_id of "1.0" breaks
        # the downstream valueUrl/multivalued string comparisons.
        from pathlib import Path

        from builder.tools.data_content import populate_condition_table

        state, exposure_id = self._exposure(tmp_path)
        book = tmp_path / "platemap.xlsx"
        self._workbook(book)
        out = populate_condition_table(state, exposure_id, str(book))
        first = Path(out["path"]).read_text(encoding="utf-8").splitlines()[1]
        assert first.startswith("1,"), first

    def test_tsv_is_not_parsed_as_one_fat_column(self, tmp_path) -> None:
        from builder.tools.data_content import populate_condition_table

        state, exposure_id = self._exposure(tmp_path)
        tsv = tmp_path / "plate.tsv"
        tsv.write_text("well_id\tcompound\nA1\tBPA\n", encoding="utf-8")
        assert populate_condition_table(state, exposure_id, str(tsv))["ok"]

    def test_unreadable_format_is_refused_by_name(self, tmp_path) -> None:
        # Previously a UnicodeDecodeError escaped into the spine's generic
        # handler; now it names the file and the reader that could not take it.
        from builder.tools.data_content import populate_condition_table

        state, exposure_id = self._exposure(tmp_path)
        sop = tmp_path / "sop.docx"
        sop.write_bytes(b"PK\x03\x04binary")
        out = populate_condition_table(state, exposure_id, str(sop))
        assert out["ok"] is False
        assert out["read_failed"] is True
        assert "sop.docx" in out["error"]
        assert ".docx" in out["reader"]

    def test_workbook_with_no_usable_sheet_says_what_it_tried(self, tmp_path) -> None:
        import openpyxl

        from builder.tools.data_content import populate_condition_table

        state, exposure_id = self._exposure(tmp_path)
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Notes"
        sheet.append(["comment"])
        sheet.append(["nothing tabular here"])
        path = tmp_path / "notes.xlsx"
        book.save(path)
        out = populate_condition_table(state, exposure_id, str(path))
        assert out["ok"] is False
        assert "no sheet" in out["error"]
        assert "Notes" in out["error"]


class TestPayloadLayerRunsOnThePipelineArm:
    """The Frictionless data-content layer is REQUIRED and used to never run (#409).

    AGENTS.md lists it as REQUIRED, but the spine only ever called
    `build_and_validate(profile="all")`, and `DATA_CONTENT_PROFILE = "data"` sits
    deliberately outside the `all|base|isa|tox` set (#95). A crate could ship a
    CSV contradicting its own declared `tableSchema` and every pass reported
    clean. These cover the invocation, not the primitive — `validate_table`
    itself is already exercised above.
    """

    @staticmethod
    def _state_with(tmp_path, compounds=("Thyroxine",), cell_lines=("CHO-K1",)):
        state = CrateState()
        state.metadata.output_path = str(tmp_path)
        for i, name in enumerate(compounds, 1):
            state.add_entity(
                Entity(
                    entity_id=f"chem_{i:03d}",
                    type="MolecularEntity",
                    fields={"name": name},
                    _provenance=EntityProvenance(created_by="llm"),
                )
            )
        for i, name in enumerate(cell_lines, 1):
            state.add_entity(
                Entity(
                    entity_id=f"cell_{i:03d}",
                    type="CellLineSample",
                    fields={"name": name},
                    _provenance=EntityProvenance(created_by="llm"),
                )
            )
        return state

    @staticmethod
    def _engine(state):
        class _Engine:
            def __init__(self, s):
                self.state = s

        return _Engine(state)

    @staticmethod
    def _write(tmp_path, rows: list[str]) -> str:
        path = tmp_path / "condition_table.csv"
        header = ",".join(c["titles"] for c in _CONDITION_TABLE_COLUMNS)
        path.write_text("\n".join([header, *rows]) + "\n", encoding="utf-8")
        return str(path)

    def test_a_bad_cell_is_reported(self, tmp_path):
        """The exact defect the missing layer let through: a non-numeric dose."""
        from builder.agents.pipeline.pipeline import _validate_populated_tables

        path = self._write(tmp_path, ["1,Uptake,CHO-K1,Thyroxine,not-a-number,uM,24h,,,"])
        issues = _validate_populated_tables(
            self._engine(self._state_with(tmp_path)),
            {"condition_table": {"populated": True, "rows": 1, "path": path}},
        )

        assert issues, "a non-numeric concentration_value must be reported"
        assert any("concentration_value" in str(i.get("property") or "") for i in issues)
        assert all(i["profile"] == "data" for i in issues)

    def test_an_unknown_compound_is_reported(self, tmp_path):
        """A cell naming no entity in the crate — the foreign-key half of the layer."""
        from builder.agents.pipeline.pipeline import _validate_populated_tables

        path = self._write(tmp_path, ["1,Uptake,CHO-K1,Mystery Compound,10,uM,24h,,,"])
        issues = _validate_populated_tables(
            self._engine(self._state_with(tmp_path)),
            {"condition_table": {"populated": True, "rows": 1, "path": path}},
        )

        assert any("Mystery Compound" in i["message"] for i in issues), issues

    def test_a_correct_table_is_clean(self, tmp_path):
        """Guards the failure mode that makes a checker worthless: firing on good data.

        The `compound`/`cell_line` cells carry entity NAMES, so an id-only
        allow-list would flag every row of a perfectly correct table.
        """
        from builder.agents.pipeline.pipeline import _validate_populated_tables

        path = self._write(
            tmp_path,
            [
                "1,Uptake,CHO-K1,Thyroxine,10,uM,24h,,,",
                "2,Uptake,CHO-K1,Silychristin,2.5,uM,24h,,,",
            ],
        )
        issues = _validate_populated_tables(
            self._engine(self._state_with(tmp_path, compounds=("Thyroxine", "Silychristin"))),
            {"condition_table": {"populated": True, "rows": 2, "path": path}},
        )

        assert issues == [], issues

    def test_a_blank_optional_cell_is_not_an_issue(self, tmp_path):
        """`propose_condition_rows` leaves a cell blank when the crate never states it.

        D5 requires that blank; flagging it would punish the honest behaviour.
        """
        from builder.agents.pipeline.pipeline import _validate_populated_tables

        path = self._write(tmp_path, ["1,Uptake,,Thyroxine,,,,,,"])
        issues = _validate_populated_tables(
            self._engine(self._state_with(tmp_path)),
            {"condition_table": {"populated": True, "rows": 1, "path": path}},
        )

        assert issues == [], issues

    def test_an_entity_id_is_accepted_as_well_as_a_name(self, tmp_path):
        """`propose_condition_rows` falls back to entity_id for an unnamed compound."""
        from builder.agents.pipeline.pipeline import _validate_populated_tables

        path = self._write(tmp_path, ["1,Uptake,CHO-K1,chem_001,10,uM,24h,,,"])
        issues = _validate_populated_tables(
            self._engine(self._state_with(tmp_path)),
            {"condition_table": {"populated": True, "rows": 1, "path": path}},
        )

        assert issues == [], issues

    @pytest.mark.parametrize(
        "materialized",
        [
            {},
            {"condition_table": {"populated": False, "reason": "no plate map"}},
            {"condition_table": {"populated": True, "rows": 0, "path": "x.csv"}},
            {"condition_table": {"populated": True, "rows": 3, "path": ""}},
        ],
        ids=["absent", "not-populated", "zero-rows", "no-path"],
    )
    def test_it_stays_silent_when_no_rows_landed(self, materialized):
        """The header-only placeholder is valid by construction; validating it is cost."""
        from builder.agents.pipeline.pipeline import _validate_populated_tables

        assert _validate_populated_tables(self._engine(CrateState()), materialized) == []

    def test_a_missing_table_never_breaks_the_build(self, tmp_path):
        """The payload layer is additive — it must not fail a build whose metadata is fine."""
        from builder.agents.pipeline.pipeline import _validate_populated_tables

        issues = _validate_populated_tables(
            self._engine(self._state_with(tmp_path)),
            {"condition_table": {"populated": True, "rows": 2, "path": str(tmp_path / "gone.csv")}},
        )

        assert issues == []

    def test_data_issues_are_not_folded_into_conformance(self, tmp_path):
        """A data-cell defect is not a SHACL failure; the two keys must stay apart.

        Collapsing them would make `success` in the eval harness mean two
        different things (#409, AGENTS.md §6).
        """
        from builder.agents.pipeline import pipeline as mod

        path = self._write(tmp_path, ["1,Uptake,CHO-K1,Thyroxine,not-a-number,uM,24h,,,"])
        issues = mod._validate_populated_tables(
            self._engine(self._state_with(tmp_path)),
            {"condition_table": {"populated": True, "rows": 1, "path": path}},
        )

        assert issues
        assert all(i["profile"] == "data" for i in issues)
        assert all(i["profile"] not in {"base", "isa", "tox"} for i in issues)
