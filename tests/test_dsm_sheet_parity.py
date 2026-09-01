"""Our DSM grid must equal the published sheet's, cell for cell.

The FAIRplus DSM's own output is a "% Complete" grid — six levels x
{Content & context, Representation & format, Hosting env. capabilities, Total} —
computed by the vendored workbook (``fair/fairplus_dsm_v1.2.xlsx``, sheet
"FAIR-DSM Assessment Sheet v1.2"). A depositor who fills that sheet in by hand, or
answers the online tool at https://fairdsm.biospeak.solutions/, must reach the numbers
this tool publishes. These tests are that guarantee.

**How this avoids being circular.** ``scripts/gen_dsm_indicators.py`` parses the sheet
into *structure* (member lists, denominators, promotion rules) which the scorer then
trusts. The evaluator below instead interprets the sheet's *formula text* directly, so
a generator that drops a repeated member, reads the wrong denominator or mis-parses a
promotion chain produces a YAML the scorer believes and this evaluator contradicts.

**What an answer vector may hold.** Three values, not two: 1, 0, and *unanswered* —
the state this tool can report and the sheet cannot. The sheet's position is that there
is no third state (column J is entirely ``=H{row}``, so a blank is numeric 0), and that
is a claim about arithmetic, so it is checked against the workbook like every other:
:class:`TestAnUnansweredIndicatorIsABlank` drives blanks through both sides. Vectors
that answer every row cannot see the difference — a scorer that drops unanswered rows
from the denominator agrees with one that counts them as 0 whenever there is nothing to
drop, which is how that bug survived this file once already (#717).

**Why a hand-rolled evaluator.** openpyxl has no calculation engine. The vendored file
does carry Excel's cached values, but only for the shipped answer vector, which is
degenerate (every Level-0 row 1, everything else 0). That one vector is still worth
asserting — it is the only arithmetic the vendor itself certifies, and it pins the
Level-0 inversion — but proving parity needs vectors Excel never evaluated. The grammar
in play is four productions, so the evaluator is twenty lines rather than a dependency.
"""

from __future__ import annotations

import pathlib
import random
import re
from collections.abc import Mapping

import openpyxl
import pytest
import yaml

from builder.state import CrateState
from builder.tools.assessment_graph import Verdict
from builder.tools.fair_assessment import _apply_promotion, dsm_grid

REPO = pathlib.Path(__file__).resolve().parents[1]
DSM_XLSX = REPO / "fair" / "fairplus_dsm_v1.2.xlsx"
DSM_YAML = REPO / "fair" / "dsm_indicators.yaml"
SHEET = "FAIR-DSM Assessment Sheet v1.2"

# Column J's IF-chain is at most three deep (J10 <- J11 <- J12).
_PLAIN = re.compile(r"^=H(\d+)$")
_PROMOTE = re.compile(r"^=IF\(J(\d+)=1,1,H(\d+)\)$")
_COUNTIFS = re.compile(r"COUNTIFS\(J(\d+),(\d)\)")
_COUNT = re.compile(r"COUNT\(([^)]*)\)")


@pytest.fixture(scope="module")
def sheet():
    return openpyxl.load_workbook(DSM_XLSX, data_only=False)[SHEET]


@pytest.fixture(scope="module")
def cached():
    return openpyxl.load_workbook(DSM_XLSX, data_only=True)[SHEET]


@pytest.fixture(scope="module")
def rows(sheet) -> dict[int, str]:
    """Sheet row -> indicator id, for every row the scoring formulas can reference."""
    out = {}
    for row in range(2, sheet.max_row + 1):
        ident = sheet.cell(row=row, column=5).value
        ident = str(ident).strip() if ident else ""
        if ident.startswith("DSM-"):
            out[row] = ident
    return out


@pytest.fixture(scope="module")
def data() -> dict:
    return yaml.safe_load(DSM_YAML.read_text())


def _evaluate(sheet, answers: dict[int, int]) -> dict[str, float]:
    """The sheet's P column, evaluated from an answer vector keyed by row."""
    validated: dict[int, int] = {}

    def j(row: int) -> int:
        if row not in validated:
            formula = str(sheet.cell(row=row, column=10).value or "").replace(" ", "")
            match = _PROMOTE.match(formula)
            validated[row] = 1 if (match and j(int(match.group(1))) == 1) else answers[row]
        return validated[row]

    out: dict[str, float] = {}
    for row in range(2, sheet.max_row + 1):
        formula = sheet.cell(row=row, column=16).value
        if formula is None or str(sheet.cell(row=row, column=15).value or "").strip() == "":
            continue
        if isinstance(formula, (int, float)):
            out[f"P{row}"] = float(formula)
            continue
        text = str(formula).replace(" ", "").lstrip("=").replace("++", "+")
        text = _COUNTIFS.sub(lambda m: f"({j(int(m.group(1)))}=={m.group(2)})", text)
        text = _COUNT.sub(lambda m: str(len(m.group(1).split(","))), text)
        out[f"P{row}"] = float(eval(text, {"__builtins__": {}}))  # noqa: S307 — fixed grammar
    return out


def _our_grid(
    data: dict, rows: dict[int, str], answers: Mapping[int, int | None]
) -> dict[str, dict]:
    """Our own arithmetic over the same answer vector, keyed by sheet cell.

    ``None`` is the state the sheet cannot hold: an indicator nothing measured. It
    reaches the scorer as ``Verdict(None)`` and reaches the sheet as the blank it is
    (:func:`_blank_is_zero`), which is the whole comparison in
    :class:`TestAnUnansweredIndicatorIsABlank`.
    """
    verdicts = {
        ident: Verdict(None if answers[row] is None else bool(answers[row]), "")
        for row, ident in rows.items()
    }
    _apply_promotion(verdicts, data["scoring"]["promotion"])
    grid = dsm_grid(CrateState(), data, None, answers=verdicts)
    return {cell["cell"]: cell for by_cat in grid.values() for cell in by_cat.values()}


def _blank_is_zero(answers: Mapping[int, int | None]) -> dict[int, int]:
    """The same vector as the workbook holds it: column J is entirely ``=H{row}``, so
    an empty H evaluates to numeric 0. There is no third value to give the evaluator."""
    return {row: 0 if value is None else value for row, value in answers.items()}


def _vector(rows: dict[int, str], fill) -> dict[int, int | None]:
    """An answer per sheet row — 1, 0, or ``None`` for unanswered.

    Rows 70 and 75 both carry DSM-4-H2 (see the test), so they answer together."""
    answers = {row: fill(row) for row in rows}
    answers[75] = answers[70]
    return answers


class TestTheWorkbooksOwnArithmetic:
    def test_the_shipped_vector_reproduces_excels_cached_values(self, sheet, cached, rows, data):
        """The one vector Excel itself evaluated: every Level-0 row 1, the rest 0.

        It is degenerate — every P cell caches 0 — but it is the vendor's own certified
        arithmetic, and it is the sharpest possible test of the Level-0 inversion: a
        scorer that counts 1s rather than 0s reads 100 where the sheet reads 0.
        """
        answers = _vector(rows, lambda row: int(sheet.cell(row=row, column=8).value or 0))
        assert sum(_blank_is_zero(answers).values()) == 11, (
            "the shipped H column is the Level-0 states"
        )

        ours = _our_grid(data, rows, answers)
        for cell_name, cell in ours.items():
            expected = cached[cell_name].value
            if expected is None:
                continue
            assert cell["published_pct"] == pytest.approx(float(expected), abs=0.05), cell_name


class TestOurGridEqualsTheSheets:
    @pytest.mark.parametrize("name", ["all-no", "all-yes"])
    def test_the_extremes(self, sheet, rows, data, name):
        answers = _vector(rows, lambda _row: 1 if name == "all-yes" else 0)
        ours = _our_grid(data, rows, answers)
        theirs = _evaluate(sheet, _blank_is_zero(answers))
        for cell_name, cell in ours.items():
            assert cell["published_pct"] == pytest.approx(theirs[cell_name], abs=0.05), cell_name

    @pytest.mark.parametrize("seed", range(5))
    def test_arbitrary_answer_vectors(self, sheet, rows, data, seed):
        rng = random.Random(seed)
        answers = _vector(rows, lambda _row: rng.randint(0, 1))
        ours = _our_grid(data, rows, answers)
        theirs = _evaluate(sheet, _blank_is_zero(answers))
        for cell_name, cell in ours.items():
            assert cell["published_pct"] == pytest.approx(theirs[cell_name], abs=0.05), cell_name

    def test_each_promotion_rule_is_exercised_on_its_own(self, sheet, rows, data):
        """One vector per rule: source met, everything else unanswered.

        A scorer that ignores promotion entirely passes every other vector in this file
        often enough to look healthy; these nine do not let it. The rest of the vector is
        genuinely unanswered rather than failed, which is what ``IF(J5=1,1,H4)`` acts on:
        the rule fires over a blank target exactly as it fires over a zero.
        """
        by_id = {ident: row for row, ident in rows.items()}
        for rule in data["scoring"]["promotion"]:
            answers = _vector(rows, lambda _row: None)
            answers[by_id[rule["when"]]] = 1
            answers[75] = answers[70]
            ours = _our_grid(data, rows, answers)
            theirs = _evaluate(sheet, _blank_is_zero(answers))
            for cell_name, cell in ours.items():
                assert cell["published_pct"] == pytest.approx(theirs[cell_name], abs=0.05), (
                    f"{rule['cell']} ({rule['when']} promotes {rule['then']}): {cell_name}"
                )


class TestTheFixturesOwnConstraints:
    """Two properties this file depends on, asserted rather than assumed."""

    def test_only_dsm_4_h2_appears_on_two_rows(self, rows):
        """Our verdict map is keyed by indicator id, so it cannot hold two answers for
        one id. The sheet lists DSM-4-H2 on rows 70 and 75 (row 75's text is the
        semantic-*search* statement, which MASTER calls DSM-4-H3). Both are hosting
        indicators no crate can evidence, so the limitation is inert today."""
        seen: dict[str, list[int]] = {}
        for row, ident in rows.items():
            seen.setdefault(ident, []).append(row)
        assert {k: v for k, v in seen.items() if len(v) > 1} == {"DSM-4-H2": [70, 75]}

    def test_an_off_sheet_indicator_moves_no_cell(self, sheet, rows, data):
        """DSM-2-R5 is scored locally but appears in no published cell, so answering it
        either way must leave every percentage untouched."""
        assert "DSM-2-R5" in data["scoring"]["off_sheet"]
        base = _vector(rows, lambda _row: 0)
        verdicts = {ident: Verdict(False, "") for ident in rows.values()}
        grid_without = _our_grid(data, rows, base)

        verdicts["DSM-2-R5"] = Verdict(True, "")
        _apply_promotion(verdicts, data["scoring"]["promotion"])
        grid_with = dsm_grid(CrateState(), data, None, answers=verdicts)
        flat = {c["cell"]: c["published_pct"] for by in grid_with.values() for c in by.values()}
        assert flat == {name: c["published_pct"] for name, c in grid_without.items()}


class TestAnUnansweredIndicatorIsABlank:
    """The state the sheet has no cell for.

    Every vector above answers every row, so none of them can tell a scorer that counts
    a blank as 0 from one that drops it out of the denominator — the two agree wherever
    there is nothing to drop. That is the bug #704 fixed, and the shape this file could
    not see (#717).

    The sheet's position is unambiguous: column J is entirely ``=H{row}`` formulas, so
    an empty H is numeric 0 and ``COUNT`` counts it. So an unanswered indicator must
    move ``published_pct`` exactly as a failed one does — and at Level 0, where the
    statements are negative and ``COUNTIFS(...,0)`` counts zeros, it must count as
    **satisfied**. Nothing else the tool publishes may follow from not having looked.
    """

    def test_nothing_answered_reads_as_the_sheet_reads_blanks(self, sheet, rows, data):
        answers = _vector(rows, lambda _row: None)
        ours = _our_grid(data, rows, answers)
        theirs = _evaluate(sheet, _blank_is_zero(answers))
        for cell_name, cell in ours.items():
            assert cell["published_pct"] == pytest.approx(theirs[cell_name], abs=0.05), cell_name

    def test_level_zero_reads_100_off_nothing_and_says_so(self, rows, data):
        """The number that made this worth pinning: a Level-0 cell nobody measured
        publishes 100% — the sheet's own arithmetic, saying the deposit escaped the
        pre-FAIRification state on the strength of never having been asked. It is only
        honest beside ``assessed``, so both are asserted together."""
        ours = _our_grid(data, rows, _vector(rows, lambda _row: None))
        zero_row = [cell for name, cell in ours.items() if cell["cell"] in {"P6", "P7", "P8", "P9"}]
        assert zero_row, "the Level-0 row is P6-P9"
        for cell in zero_row:
            assert cell["published_pct"] == 100.0, cell["cell"]
            assert cell["assessed"] == 0, cell["cell"]
            assert cell["pct"] is None, cell["cell"]

    def test_an_unanswered_row_scores_like_a_failed_one(self, sheet, rows, data):
        """Not a claim about our code — a claim about the sheet, checked against it.
        The same vector with blanks and with zeros must produce the same grid."""
        blanks = _vector(rows, lambda row: None if row % 3 else 1)
        zeros = {row: 0 if value is None else value for row, value in blanks.items()}
        ours_blank = _our_grid(data, rows, blanks)
        ours_zero = _our_grid(data, rows, zeros)
        theirs = _evaluate(sheet, zeros)
        for cell_name, cell in ours_blank.items():
            assert cell["published_pct"] == ours_zero[cell_name]["published_pct"], cell_name
            assert cell["published_pct"] == pytest.approx(theirs[cell_name], abs=0.05), cell_name

    @pytest.mark.parametrize("seed", range(3))
    def test_the_shape_a_real_crate_produces(self, sheet, rows, data, seed):
        """The production vector: the 39 indicators no crate can evidence are `na` and
        stay unanswered unless a depositor supplies them, so nearly half of every real
        assessment is blank. That is the vector parity most needs to hold on."""
        rng = random.Random(seed)
        na = {i["id"] for i in data["indicators"] if i.get("scope") == "na"}
        assert len(na) > 30, "the unanswered half is the point of this test"
        answers = _vector(rows, lambda row: None if rows[row] in na else rng.randint(0, 1))
        ours = _our_grid(data, rows, answers)
        theirs = _evaluate(sheet, _blank_is_zero(answers))
        for cell_name, cell in ours.items():
            assert cell["published_pct"] == pytest.approx(theirs[cell_name], abs=0.05), cell_name
