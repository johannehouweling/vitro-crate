"""Format-aware file readers for structured and unstructured documents.

Provides dedicated readers for Excel (``.xlsx``), Word (``.docx``), and a
unified ``read_file`` dispatcher that routes to the right reader based on
file extension.  All readers enforce size and row limits so the LLM never
gets flooded with large files.
"""

from __future__ import annotations

import json
import logging
import re
import warnings
from pathlib import Path
from typing import Any

import yaml

logger = logging.getLogger(__name__)


def _silence_openpyxl_extension_warnings() -> None:
    """Stop openpyxl reporting the workbook features it drops while reading.

    openpyxl warns once per worksheet extension it cannot round-trip —
    ``"Data Validation extension is not supported and will be removed"``, and the
    same sentence for Conditional Formatting, Sparkline Group, Slicer List,
    Protected Range, Ignored Error, Web Extension and Timeline Ref
    (``openpyxl.xml.constants.EXT_TYPES``).

    Every one of those is an Excel *presentation* feature: a dropdown, a colour
    rule, a slicer. We open workbooks ``read_only=True, data_only=True`` to take
    cell VALUES, and we never write a workbook back — so nothing openpyxl drops
    can change what we read, and "will be removed" describes openpyxl's in-memory
    model, not the file on disk. A depositor workbook with validated columns
    emitted pages of it, and neither the user nor the agent can act on any of it.

    Registered at import rather than around each ``load_workbook``:
    :func:`warnings.catch_warnings` mutates process-global state and is not
    thread-safe, and these readers run concurrently in the tools node.

    Matched on the message alone. The ``module`` argument matches the module that
    *issues* the warning, which openpyxl reaches through a ``warn()`` helper — a
    detail of theirs we would rather not depend on. The sentence is distinctive
    enough on its own, and the filter is pinned to UserWarning.
    """
    warnings.filterwarnings(
        "ignore",
        message=r".*extension is not supported and will be removed",
        category=UserWarning,
    )


_silence_openpyxl_extension_warnings()

# ---------------------------------------------------------------------------
# Limits
# ---------------------------------------------------------------------------

# Unified size ceiling (Issue #148): matches scanner.read_file_sample /
# extract_pdf_text (100 MB). The readers cap rows/lines independently, so memory
# stays bounded even for large files; the old 1 MB ceiling silently returned
# None for ordinary mid-size files (e.g. a 5 MB CSV), starving the agent.
_MAX_BYTES = 100 * 1024 * 1024  # 100 MB — skip files larger than this
_MAX_ROWS = 500  # max rows to return from structured formats

# Full-return byte budget for plain-text / JSON content (Issue #240). A 32 KB
# JSON is only ~8K tokens and must come back COMPLETE — the old 100-line cap
# dropped the tail, so a weak model never saw fields deep in the file and looped
# "let me read the rest". We return text in full up to this budget; only a file
# that genuinely exceeds it is truncated, and then with an explicit marker.
_TEXT_BUDGET_BYTES = 64 * 1024  # 64 KiB — generous full-return budget for text


def _format_kib(num_bytes: int) -> str:
    """Format a byte count as a compact ``N.N KiB`` string for messages."""
    return f"{num_bytes / 1024:.1f} KiB"


# How many concrete child-file paths to surface in a directory message. Enough
# for a weak model to pick a real file to read next, capped so the tool message
# stays compact on a large directory.
_DIR_LISTING_LIMIT = 25


def _directory_message(path: str) -> str:
    """Actionable message for a reader that was handed a directory (Issue #240).

    The LLM kept calling ``read_file``/``read_file_sample`` on a *directory* and
    got a silent ``None`` each time, then looped. The abstract "use
    list_scanned_files" hint alone still looped a weak model, so this also lists
    the directory's immediate **readable file children** as CONCRETE paths the
    model can read next — the most direct way to break the loop (follow-up to
    #240). Subdirectories are excluded (they would just loop the same way); an
    empty/unreadable directory falls back to the plain guidance.
    """
    base = (
        f"{path} is a directory, not a file — use list_scanned_files to browse "
        f"the inventory, then read a specific file by its path."
    )
    try:
        children = sorted(
            entry
            for entry in Path(path).iterdir()
            if entry.is_file() and not entry.name.startswith(".")
        )
    except OSError:
        return base
    if not children:
        return base

    shown = children[:_DIR_LISTING_LIMIT]
    listed = "\n".join(f"  - {child}" for child in shown)
    more = ""
    if len(children) > len(shown):
        more = f"\n  …and {len(children) - len(shown)} more (use list_scanned_files)."
    return (
        f"{base}\nReadable files in this directory — read one of these paths "
        f"directly:\n{listed}{more}"
    )


# ---------------------------------------------------------------------------
# Excel (.xlsx)
# ---------------------------------------------------------------------------


def read_excel_rows(
    path: str,
    *,
    max_rows: int | None = None,
    max_bytes: int = _MAX_BYTES,
) -> dict[str, list[dict[str, Any]]] | None:
    """Read an ``.xlsx`` workbook as TYPED rows, one list of dicts per sheet.

    Distinct from :func:`read_excel`, which renders pipe-delimited *text* for a
    model to read: re-parsing that as CSV loses cell types and breaks on any
    value containing a pipe. A plate map has to round-trip as data, so this
    returns the cells.

    ``max_rows`` defaults to **no cap** deliberately. The 500-row default that
    suits an LLM preview would silently drop wells from a 384- or 1536-well
    plate — the same silent-data-loss class as the bug this exists to fix. A
    caller that does pass a cap gets a ``__truncated__`` key so it can say so.

    Values are normalised at this boundary: ``None`` becomes ``""`` and an
    integral float becomes an ``int``, so a ``well_id`` reads ``1`` rather than
    ``1.0`` — openpyxl types every numeric cell as float, and the downstream
    valueUrl/multivalued reasoning compares strings.

    Returns ``None`` when the file is missing, too large, or unreadable.
    """
    src = Path(path)
    if not src.is_file():
        logger.warning("read_excel_rows: not a file: %s", path)
        return None
    if src.stat().st_size > max_bytes:
        logger.warning("read_excel_rows: %s exceeds %d bytes — skipped", path, max_bytes)
        return None
    try:
        import openpyxl
    except ImportError:
        logger.error("read_excel_rows: openpyxl is not installed")
        return None
    try:
        book = openpyxl.load_workbook(src, read_only=True, data_only=True, keep_links=False)
    except Exception:
        logger.exception("read_excel_rows: could not open %s", path)
        return None

    def _cell(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value if isinstance(value, (int, str)) else str(value)

    sheets: dict[str, list[dict[str, Any]]] = {}
    try:
        for sheet in book.worksheets:
            header: list[str] | None = None
            rows: list[dict[str, Any]] = []
            for raw in sheet.iter_rows(values_only=True):
                cells = [_cell(v) for v in raw]
                if not any(str(c).strip() for c in cells):
                    continue
                if header is None:
                    header = [str(c).strip() for c in cells]
                    continue
                row = {
                    key: cells[i] if i < len(cells) else "" for i, key in enumerate(header) if key
                }
                rows.append(row)
                if max_rows is not None and len(rows) >= max_rows:
                    rows.append({"__truncated__": True})
                    break
            sheets[sheet.title] = rows
    finally:
        book.close()
    return sheets


# openpyxl reads OOXML only, so every pre-2007 ``.xls`` raised
# InvalidFileException and was logged as a full ERROR traceback — once per file,
# and the real deposit corpus is full of them (#417). BIFF needs its own reader.
_OOXML_SUFFIXES: frozenset[str] = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})


def _read_xls_biff(file_path: Path, max_rows: int) -> str | None:
    """Read a legacy BIFF ``.xls`` workbook into the same pipe-delimited text.

    Byte-identical formatting to the openpyxl branch so a caller cannot tell
    which reader ran.
    """
    try:
        import xlrd
    except ImportError:
        logger.warning(
            "Cannot read legacy .xls %s: xlrd is not installed "
            "(pip install xlrd) — the file contributes nothing to the crate",
            file_path.name,
        )
        return None

    import io

    try:
        # logfile is REQUIRED: xlrd defaults to sys.stdout and prints codepage
        # notices straight into the terminal. Replacing a traceback with stdout
        # spew is not a fix.
        book = xlrd.open_workbook(str(file_path), logfile=io.StringIO(), on_demand=True)
    except Exception as exc:  # noqa: BLE001 — an unreadable file is not fatal
        logger.warning(
            "Cannot read legacy .xls %s: %s — the file contributes nothing to the crate",
            file_path.name,
            exc,
        )
        logger.debug("Legacy .xls read failed for %s", file_path, exc_info=True)
        return None

    def _cell(cell: Any, datemode: int) -> str:
        if cell.ctype == xlrd.XL_CELL_EMPTY or cell.value is None:
            return ""
        if cell.ctype == xlrd.XL_CELL_DATE:
            # xlrd hands back a raw serial float; without this the drafter reads
            # 44637.0 where the sheet says 2022-03-17 — corrupted data in the
            # crate, which D5 forbids.
            try:
                return str(xlrd.xldate.xldate_as_datetime(cell.value, datemode))
            except Exception:  # noqa: BLE001 — a bad serial is not worth failing on
                return str(cell.value)
        value = cell.value
        if isinstance(value, float) and value.is_integer():
            return str(int(value))  # 7, not 7.0 — matches the openpyxl branch
        return str(value)

    parts: list[str] = []
    try:
        for sheet in book.sheets():
            parts.append(f"[Sheet: {sheet.name}]")
            row_count = 0
            for index in range(sheet.nrows):
                cells = [_cell(c, book.datemode) for c in sheet.row(index)]
                if not any(c for c in cells):
                    continue
                parts.append("| " + " | ".join(cells) + " |")
                row_count += 1
                if row_count >= max_rows:
                    parts.append(f"[... truncated at {max_rows} rows]")
                    break
            parts.append("")
    except Exception:  # noqa: BLE001 — keep whatever rows were already read
        logger.warning("Legacy .xls %s ended early (corrupt sheet)", file_path.name)
        logger.debug("Row extraction failed for %s", file_path, exc_info=True)
    finally:
        book.release_resources()

    return "\n".join(parts).rstrip("\n")


def read_excel(
    path: str,
    *,
    max_rows: int = _MAX_ROWS,
    max_bytes: int = _MAX_BYTES,
    compact: bool = False,
) -> str | None:
    """Read an Excel ``.xlsx`` file and return its content as pipe-delimited text.

    Each worksheet is returned with a ``[Sheet: <name>]`` header followed by
    pipe-delimited rows.  Empty rows are skipped.  Returns *None* if the file
    cannot be read, is too large, or is not a valid Excel workbook.

    Args:
        path: Path to the ``.xlsx`` file.
        max_rows: Maximum number of data rows to return per sheet (default 500).
        max_bytes: Maximum file size in bytes (default 1 MB).  Larger files
            are skipped.
        compact: Run the result through :func:`compact_grid_text` (#378), which
            strips the repeated header row, the Comments column and empty cells.
            Defaults to **False** so existing callers see byte-identical output.

    Returns:
        Pipe-delimited text with sheet headers, or *None*.
    """
    file_path = Path(path)
    if not file_path.is_file():
        return None

    try:
        size = file_path.stat().st_size
        if size > max_bytes:
            logger.info("Skipping large Excel file (>%d bytes): %s", max_bytes, path)
            return None
    except OSError:
        return None

    # Legacy BIFF is a different container entirely — openpyxl cannot open it.
    if file_path.suffix.lower() not in _OOXML_SUFFIXES:
        text = _read_xls_biff(file_path, max_rows)
        if text is None:
            return None
        return compact_grid_text(text) if compact else text

    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is not installed.")
        return None

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:  # noqa: BLE001
        # One warning naming the reason, not a traceback per file (#417): a
        # corpus of legacy workbooks produced pages of stack traces that read
        # like a crash while the scan was in fact fine.
        logger.warning(
            "Cannot read Excel file %s: %s — the file contributes nothing to the crate",
            file_path.name,
            exc,
        )
        logger.debug("Excel open failed for %s", path, exc_info=True)
        return None

    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"[Sheet: {sheet_name}]")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if all(cell is None for cell in row):
                    continue
                cleaned = [str(cell) if cell is not None else "" for cell in row]
                parts.append("| " + " | ".join(cleaned) + " |")
                row_count += 1
                if row_count >= max_rows:
                    parts.append(f"[... truncated at {max_rows} rows]")
                    break
            parts.append("")
    finally:
        wb.close()

    text = "\n".join(parts).rstrip("\n")
    return compact_grid_text(text) if compact else text


# ---------------------------------------------------------------------------
# Word (.docx)
# ---------------------------------------------------------------------------


def read_docx(
    path: str,
    *,
    max_bytes: int = _MAX_BYTES,
) -> str | None:
    """Read a Word ``.docx`` file and return its text content.

    Extracts paragraph text, table content (as pipe-delimited rows), and
    section headings where available.  Returns *None* if the file cannot be
    read, is too large, or is not a valid ``.docx``.

    Args:
        path: Path to the ``.docx`` file.
        max_bytes: Maximum file size in bytes (default 1 MB).

    Returns:
        Plain text with ``[Table N]`` markers, or *None*.
    """
    file_path = Path(path)
    if not file_path.is_file():
        return None

    try:
        size = file_path.stat().st_size
        if size > max_bytes:
            logger.info("Skipping large DOCX file (>%d bytes): %s", max_bytes, path)
            return None
    except OSError:
        return None

    try:
        from docx import Document  # type: ignore[import-untyped]
    except ImportError:
        logger.error("python-docx is not installed.")
        return None

    try:
        doc = Document(str(file_path))
    except Exception:
        logger.exception("Error opening DOCX file: %s", path)
        return None

    parts: list[str] = []
    table_idx = 0

    body = doc.element.body
    for child in body:
        tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag

        if tag == "p":
            from docx.text.paragraph import Paragraph

            para = Paragraph(child, doc)
            text = para.text.strip()
            if text:
                parts.append(text)

        elif tag == "tbl":
            from docx.table import Table as DocxTable

            table = DocxTable(child, doc)
            table_idx += 1
            table_rows: list[str] = []

            for row in table.rows:
                cells = [cell.text.strip() for cell in row.cells]
                if any(cells):
                    table_rows.append("| " + " | ".join(cells) + " |")

            if table_rows:
                parts.append(f"[Table {table_idx} ({len(table_rows)} rows)]")
                parts.append(table_rows[0])
                parts.append("| " + " | ".join("---" for _ in table.rows[0].cells) + " |")
                parts.extend(table_rows[1:])

    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Unified dispatcher
# ---------------------------------------------------------------------------


def _read_text_file(
    path: str,
    *,
    max_bytes: int = _MAX_BYTES,
    budget: int | None = None,
) -> str | None:
    """Read a plain-text file IN FULL, up to a generous byte *budget* (Issue #240).

    Returns the whole file when it fits in ``budget`` bytes (default
    :data:`_TEXT_BUDGET_BYTES`, 64 KiB). When it exceeds the budget the content
    shown is returned PLUS an explicit, unmistakable truncation marker stating
    how much was shown and that re-reading the same way will NOT return more — so
    a weak model stops looping instead of asking for "the rest".

    The hard ``max_bytes`` safety cap (default 100 MB) still skips genuinely huge
    files entirely (returns *None*) so we never load a 16 MB binary into memory.

    Returns *None* if the file is too large (> ``max_bytes``), binary, or
    unreadable.
    """
    # Resolve the budget at call time so tests can monkeypatch the module attr.
    if budget is None:
        budget = _TEXT_BUDGET_BYTES

    file_path = Path(path)
    try:
        size = file_path.stat().st_size
        if size > max_bytes:
            logger.info("Skipping large text file (>%d bytes): %s", max_bytes, path)
            return None
    except OSError:
        return None

    try:
        with file_path.open("rb") as fb:
            if b"\x00" in fb.read(8192):
                return None
    except OSError:
        return None

    try:
        # Read up to budget+1 bytes so we can tell whether the file overflowed
        # the budget without slurping the whole (possibly large) file.
        with file_path.open("r", encoding="utf-8", errors="replace") as f:
            shown = f.read(budget)
            overflowed = f.read(1) != ""
    except PermissionError:
        logger.warning("Permission denied reading file: %s", path)
        return None
    except Exception:
        logger.exception("Error reading text file: %s", path)
        return None

    content = shown.rstrip("\n")
    if not overflowed:
        return content

    shown_bytes = len(shown.encode("utf-8"))
    marker = (
        f"\n[truncated: showing first {_format_kib(shown_bytes)} of "
        f"{_format_kib(size)}; this is the maximum for this tool — do not "
        f"re-read]"
    )
    return content + marker


# Header words that name an *authoring-guidance* column rather than a data one.
# Matched against the header cell only, never a value (#421). English and Dutch
# both appear in the real corpus: S-VHPS22's top-level file is an RIVM template
# whose six columns are Veldnaam | Optionaliteit | Hoe vaak in te vullen |
# Beschrijving | Tips | Hier invullen — four of those six are instructions to
# the depositor, carrying 9,913 chars to deliver 610 chars of actual metadata.
_GUIDANCE_HEADERS: frozenset[str] = frozenset(
    {
        "comments",
        "commentaar",
        "tips",
        "beschrijving",
        "toelichting",
        "description",
        "instructions",
        "instructie",
        "example",
        "voorbeeld",
        "optionaliteit",
        "hoe vaak in te vullen",
    }
)


def _normalize_header(cell: str) -> str:
    """Fold a header cell for guidance-vocabulary matching."""
    return " ".join((cell or "").split()).casefold().rstrip(":*")


def _guidance_columns(header: list[str]) -> set[int]:
    """Indices of the header's authoring-guidance columns."""
    return {i for i, cell in enumerate(header) if _normalize_header(cell) in _GUIDANCE_HEADERS}


def _is_guidance_cell(index: int, guidance: set[int], overflow_from: int | None) -> bool:
    """True when cell *index* belongs to a guidance column, including overflow.

    Rows are ragged in the real corpus: the S-VHPS26 sheets declare a four-column
    ``Parameter | Standard or ontology reference | Value | Comments`` header but
    emit five cells, because a long Comments entry spills into an unheadered
    column (``"AOP title and ID, e.g., …"`` then ``"URL: https://aopwiki.org/…"``
    — one instruction split in two). When the guidance column is the header's
    last, everything at or past it is that same guidance, so the tail goes too.
    A guidance column in the *middle* gets no such treatment: there, a trailing
    unheadered cell sits beyond a real data column and is not ours to judge.
    """
    return index in guidance or (overflow_from is not None and index >= overflow_from)


def _guidance_is_droppable(
    rows: list[list[str]], guidance: set[int], overflow_from: int | None
) -> bool:
    """True when the sheet still says something without its guidance columns.

    The header vocabulary alone is not safe to act on: ``Description`` names
    scaffolding in a fill-in-the-blanks template but real content in a workbook
    that simply has a description column. The two are distinguishable at the
    *sheet* level — a template has a separate answer column that the depositor
    actually filled, so at least one row stands on its own with two non-empty
    non-guidance cells. A workbook whose only content lives in ``Description``
    has none, and keeps every column.

    Deciding once per sheet rather than per row matters: a per-row test would
    keep the full instruction text on exactly the unfilled rows that carry no
    data at all, which is most of the noise being removed.
    """
    return any(
        sum(
            1
            for i, cell in enumerate(cells)
            if cell and not _is_guidance_cell(i, guidance, overflow_from)
        )
        >= 2
        for cells in rows
    )


# Guidance columns a candidate row must name before it may overrule row 0.
#
# Measured, not guessed. Across 1,046 sheets of the real corpus (220 readable
# workbooks: the three S-VHPS deposits plus the committed fixtures) there are
# 79,146 pipe rows. **Exactly two of them name two or more guidance columns —
# and both are the same RIVM header**, ``Veldnaam | Optionaliteit | Hoe vaak in
# te vullen | Beschrijving | Tips | Hier invullen``, which names four. 140 rows
# name exactly one; 79,004 name none.
#
# One is nowhere near enough: at a threshold of 1 the search re-picks on 16
# sheets across 11 workbooks, every one of them a mis-pick. 2, 3 and 4 all
# re-pick on the RIVM sheet and nothing else, because the corpus has NOTHING
# between one guidance word and that header's four — so anything in 2..4 costs
# the same on real data and the only question is which is hardest to trip by
# accident.
#
# 3, therefore: the middle of the empty gap rather than its near edge. The
# difference is not hypothetical. Two rows carrying two vocabulary words each is
# a shape this domain really produces — a legend or a DataCite crosswalk that
# TABULATES column names as data ("Beschrijving | Toelichting | 1.0 | RIVM") —
# and under a narrow title row such a row is the first spanning row, so it would
# be promoted, deleted, and its column struck from every row below it. At 2 that
# sheet is destroyed; at 3 it is left alone, and the RIVM header still wins with
# four.
#
# This is what makes the re-pick safe at all. A key/value data row —
# ``Toelichting | zie protocol | RIVM`` — carries at most one vocabulary word,
# so it can never be promoted and have its label deleted as scaffolding.
_MIN_GUIDANCE_COLUMNS = 3


def _table_span(rows: list[list[str]]) -> int:
    """How many columns this sheet's rows actually occupy.

    The median row width, deliberately: a value containing a literal ``|`` splits
    into extra cells at this layer — the RIVM sheet tells depositors to separate
    repeated values with a pipe and then does so itself — which makes the widest
    row a lie, while the mode is a coin-flip whenever two widths tie.

    Measured on the real S-VHPS22 template, whose table is six columns wide: the
    median row is 5 cells, but its ``Trefwoorden`` row lists ten pipe-separated
    keywords and arrives here as **15**. A max-width span therefore demands 15
    cells of the header, no row can reach that, and the search finds nothing —
    the file goes back to the 18,777 chars of #421.
    """
    widths = sorted(sum(1 for cell in cells if cell) for cells in rows)
    return widths[len(widths) // 2]


def _header_index(rows: list[list[str]]) -> int:
    """Index of the sheet's header among its pipe rows — 0 unless prose precedes it.

    S-VHPS22's top-level RIVM template opens with six rows of Dutch instructions
    on how to fill the sheet in, and only then names its columns (``Veldnaam |
    Optionaliteit | Hoe vaak in te vullen | Beschrijving | Tips | Hier
    invullen``). Reading the first pipe row as the header matched the guidance
    vocabulary against a sentence, so not one guidance column was recognised and
    the file compacted to 18,777 chars against a 9,000-char tier-0 share (#421).

    Row 0 keeps the job unless a lower row **earns** it, and only one shape can:
    a fill-in-the-blanks template whose header names at least
    :data:`_MIN_GUIDANCE_COLUMNS` authoring-guidance columns, sitting under
    preamble that does not reach across the table. Two things are asked of the
    candidate and nothing else is enough — it spans the table
    (:func:`_table_span`), and it names two guidance columns.

    Only the *first* spanning row is ever considered. Once a row reaches across
    the table the preamble is over; the rows after it are data, not further
    chances, and a sheet whose first spanning row is not a guidance header is
    left exactly as it was.

    Refusing is the safe outcome because a wrong pick loses data silently:
    :func:`_compact_sheet` deletes the row it calls the header and reads the
    guidance vocabulary off it, so promoting a data row would delete a real
    column from every row below it.
    """
    if not rows:
        return 0
    span = _table_span(rows)
    for idx, cells in enumerate(rows):
        if sum(1 for cell in cells if cell) < span:
            continue  # still preamble: this row does not reach across the table
        if idx and len(_guidance_columns(cells)) >= _MIN_GUIDANCE_COLUMNS:
            return idx
        # The first spanning row is the only candidate there is. If it does not
        # name guidance columns, the row after it is a data row, not a second
        # chance — keep row 0 and change nothing about this sheet.
        return 0
    return 0


def _compact_sheet(lines: list[str]) -> list[str]:
    """Apply the row/column rules to one sheet's worth of lines."""
    rows: dict[int, list[str]] = {}
    for idx, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith("|"):
            rows[idx] = [c.strip() for c in stripped.strip("|").split("|")]
    if not rows:
        return list(lines)

    ordered = list(rows.items())
    position = _header_index([cells for _, cells in ordered])
    header_idx, header_cells = ordered[position]
    guidance = _guidance_columns(header_cells)
    last = len(header_cells) - 1
    overflow_from = last if last in guidance else None
    # A header speaks for the table under it and for nothing above it. Judging a
    # preamble row by the header's column meanings deletes real content: the RIVM
    # sheet's ``Versienummer | 1.2.0 | 2025-06-26`` sits in the columns the header
    # later calls Optionaliteit and Hoe vaak in te vullen, and would lose two of
    # its three cells — then the whole row, for falling under two.
    body = [cells for idx, cells in ordered if idx > header_idx]
    if guidance and not _guidance_is_droppable(body, guidance, overflow_from):
        guidance, overflow_from = set(), None
    # ``Parameter`` in the first cell marks the repeated boilerplate header of a
    # depositor workbook, which is why that row goes even when no guidance column
    # was recognised. It is only ever safe on **row 0**: a searched row starting
    # with ``Parameter`` is a row this function inferred, and deleting it outright
    # — with no guidance column involved, so none of `_guidance_is_droppable`'s
    # protection — would silently destroy a data row under a title row.
    drop_header = bool(guidance) or (
        position == 0 and bool(header_cells) and header_cells[0].lower() == "parameter"
    )

    out: list[str] = []
    seen_iris: set[str] = set()
    for idx, line in enumerate(lines):
        cells = rows.get(idx)
        if cells is None:
            out.append(line)
            continue
        if idx == header_idx and drop_header:
            continue
        # Dropped by index, not by position: a ragged row shorter than the
        # header would otherwise lose whichever column happened to be last. Only
        # rows the header actually describes — itself and everything below it —
        # are read through its columns (see the ``body`` note above).
        in_table = idx >= header_idx
        kept = [
            c
            for i, c in enumerate(cells)
            if c and not (in_table and _is_guidance_cell(i, guidance, overflow_from))
        ]
        if len(kept) < 2:
            continue
        # An ontology IRI annotating a column repeats verbatim on every row it
        # types — four of them account for 4,133 chars on the real S-VHPS26
        # workbook. State each once per sheet (#419). Never at the cost of the
        # row's last identifying cell: a row that would fall below two cells
        # keeps its IRI, so no label is ever traded away for the saving.
        deduped = [c for c in kept if not (c in seen_iris and _IRI_CELL.match(c))]
        if len(deduped) >= 2:
            kept = deduped
        seen_iris.update(c for c in kept if _IRI_CELL.match(c))
        out.append("| " + " | ".join(kept) + " |")

    return out


def compact_grid_text(text: str) -> str:
    """Densify ``[Sheet: …]`` + pipe-row output, keeping every signal cell (#378).

    A depositor-filled metadata workbook is mostly boilerplate: a ``Parameter |
    Standard or ontology reference | Value | Comments`` header repeated per
    sheet, columns of authoring instructions, and empty cells. On the real
    S-VHPS26 workbook that noise pushes the cell line, RRID, author and
    chemicals 2-5 past any affordable context slice.

    Every column rule is read off the sheet's header, which is **found, not
    assumed to be row 1** (:func:`_header_index`): an RIVM template spends six
    rows telling the depositor how to fill the sheet in before naming a single
    column, and matching the guidance vocabulary against that prose recognised
    nothing (#421).

    Four rules, in order: drop the repeated header row; drop the sheet's
    authoring-guidance columns **only when that sheet's own header names them**
    (:data:`_GUIDANCE_HEADERS`) *and* the sheet survives without them
    (:func:`_guidance_is_droppable`); drop empty cells, then drop rows left with
    fewer than two non-empty cells; finally fold a numbered series into one row
    (see :func:`_collapse_numbered_series`).

    **Rows are never dropped on the emptiness of one named column.** That rule
    looks right and destroys the General information sheet, because this
    depositor filled column 2 rather than the ``Value`` column — so ``Dr. Fabian
    Wagenaars``, the ORCID, the DOI and the assay name all sit on rows whose
    ``Value`` cell is blank. Text carrying no pipe rows is returned unchanged.

    Header detection, guidance columns and IRI dedup are all **per sheet**, not
    per workbook: a five-sheet template mixes layouts, and a decision taken on
    sheet 1 has no authority over sheet 4.
    """
    lines = text.split("\n")

    # Split on sheet boundaries first — the column rules need to see a whole
    # sheet before deciding, which a single streaming pass cannot do.
    blocks: list[list[str]] = [[]]
    for line in lines:
        if line.strip().startswith("[Sheet:"):
            blocks.append([])
        blocks[-1].append(line)

    out: list[str] = []
    for block in blocks:
        out.extend(_compact_sheet(block))

    return "\n".join(_collapse_numbered_series(out)).strip()


# A numbered series must be at least this long to be folded. Two consecutive rows
# that merely happen to end in 1 and 2 are far more likely to be distinct
# parameters than a series, and folding them would lose a label.
_MIN_SERIES_RUN = 3

_SERIES_KEY = re.compile(r"^(?P<prefix>.+)_(?P<index>\d+)$")

# A cell that is nothing but an ontology/identifier IRI — the annotation form a
# depositor workbook repeats per row. Deliberately anchored and whole-cell: a cell
# that merely CONTAINS a URL alongside prose is real content and is never deduped.
_IRI_CELL = re.compile(r"^https?://\S+$")


def _collapse_numbered_series(rows: list[str]) -> list[str]:
    """Fold ``<prefix>_1 … <prefix>_N`` rows that differ only in their value (#419).

    A depositor-filled workbook states a dose series one row per level, repeating
    the same ontology IRI every time::

        | Chemical_1_Concentration_1 | http://nmrML.org/nmrCV#NMR:1000095 | 0.003 |
        | Chemical_1_Concentration_2 | http://nmrML.org/nmrCV#NMR:1000095 | 0.01  |

    On the real S-VHPS26 workbook 160 such rows carry 11,693 characters — 53% of
    the whole compacted sheet — to express 19 dose series. Folding each run to one
    row is what makes the full chemical table affordable inside the existing
    context budget, instead of buying it with a budget increase that would cost
    tokens on every run.

    A run folds only when the rows are CONSECUTIVE, share a prefix, number
    contiguously, agree on every cell but the last, and are at least
    :data:`_MIN_SERIES_RUN` long. A run that qualifies structurally is still
    REFUSED when folding it would not round-trip:

    * **a value containing the join separator.** ``1,2-dichloroethane`` and the
      decimal comma an EU depositor may well type (``0,03``) both make the joined
      list unreadable — the reader cannot recover how many members there were,
      let alone which index each belongs to.
    * **a value that is a bare ontology IRI.** The dedup pass above removes an
      annotation cell, so a row whose value was blank collapses to
      ``[label, IRI]`` and looks the same shape as its neighbours; folding then
      presents the IRI *as a dose*. Refusing the run keeps the rows honest.
    * **inconsistent zero-padding** across the run's indices, which a numeric
      range label cannot express.

    Rows are re-emitted VERBATIM whenever a run does not fold, so the index token
    is never rewritten — ``Aliquot_007`` must not come back as ``Aliquot_7`` — and
    a sheet without a series is returned byte-identical.
    """
    out: list[str] = []
    # (prefix, numeric index, raw index token, cells after the label, original row)
    run: list[tuple[str, int, str, list[str], str]] = []

    def _foldable() -> bool:
        if len(run) < _MIN_SERIES_RUN:
            return False
        values = [cells[-1] for *_rest, cells, _row in run]
        if any("," in v for v in values):
            return False
        if any(_IRI_CELL.match(v) for v in values):
            return False
        widths = {len(token) for _p, _i, token, _c, _r in run}
        return len(widths) == 1 or not any(t.startswith("0") for _p, _i, t, _c, _r in run)

    def _flush() -> None:
        if not run:
            return
        if not _foldable():
            out.extend(row for *_rest, row in run)
        else:
            prefix = run[0][0]
            shared = run[0][3][:-1]
            values = ", ".join(cells[-1] for _p, _i, _t, cells, _r in run)
            label = f"{prefix}_{run[0][2]}-{run[-1][2]}"
            out.append("| " + " | ".join([label, *shared, values]) + " |")
        run.clear()

    for row in rows:
        stripped = row.strip()
        cells = (
            [c.strip() for c in stripped.strip("|").split("|")] if stripped.startswith("|") else []
        )
        match = _SERIES_KEY.match(cells[0]) if len(cells) >= 2 else None
        if match is None:
            _flush()
            out.append(row)
            continue

        prefix, token, rest = match["prefix"], match["index"], cells[1:]
        index = int(token)
        contiguous = run and run[-1][0] == prefix and run[-1][1] == index - 1
        same_shape = run and run[-1][3][:-1] == rest[:-1]
        if not (contiguous and same_shape):
            _flush()
        run.append((prefix, index, token, rest, row))

    _flush()
    return out


def _flatten_attributes(attrs: Any, out: list[str], indent: str = "") -> None:
    """Emit ``name=value [Qual=Val; …]`` lines for a BioStudies attribute list."""
    for attr in attrs or []:
        if not isinstance(attr, dict):
            continue
        name = str(attr.get("name") or "").strip()
        value = str(attr.get("value") or "").strip()
        if not name and not value:
            continue
        line = f"{indent}{name}={value}" if name else f"{indent}{value}"
        quals: list[str] = []
        for qual in attr.get("valqual") or []:
            if not isinstance(qual, dict):
                continue
            qname = str(qual.get("name") or "").strip()
            qvalue = str(qual.get("value") or "").strip()
            if qname or qvalue:
                quals.append(f"{qname}={qvalue}" if qname else qvalue)
        if quals:
            line += " [" + "; ".join(quals) + "]"
        out.append(line)


def _flatten_section(node: Any, out: list[str]) -> None:
    """Walk a BioStudies section tree, emitting attributes, links and children."""
    if isinstance(node, list):
        for item in node:
            _flatten_section(item, out)
        return
    if not isinstance(node, dict):
        return

    stype = str(node.get("type") or "").strip()
    accno = str(node.get("accno") or "").strip()
    if stype or accno:
        out.append(f"[{stype or 'section'}{' ' + accno if accno else ''}]")

    _flatten_attributes(node.get("attributes"), out)

    for link in node.get("links") or []:
        if not isinstance(link, dict):
            continue
        url = str(link.get("url") or "").strip()
        if url:
            out.append(f"link={url}")
        _flatten_attributes(link.get("attributes"), out, indent="  ")

    if node.get("section") is not None:
        _flatten_section(node["section"], out)
    for sub in node.get("subsections") or []:
        _flatten_section(sub, out)


def compact_attribute_json(text: str) -> str:
    """Flatten a BioStudies ``{name, value, valqual[]}`` tree to dense lines (#378).

    A pretty-printed submission descriptor spends most of its bytes on JSON
    punctuation and indentation, so a bounded slice of the raw file stops inside
    the first section — on the real S-VHPS26 descriptor the licence, the AOP URL
    and every author sit past 2,900 characters and never reach the leaf.

    ``valqual`` entries are preserved as a bracketed suffix. Dropping them is the
    tempting simplification and it loses exactly the qualified values worth
    having — the AOP wiki URL and the BAO ontology ids.

    Input that is not JSON, or is JSON without a BioStudies attribute tree, is
    returned **unchanged** so the caller can apply this blindly.
    """
    try:
        parsed = json.loads(text)
    except (ValueError, TypeError):
        return text
    if not isinstance(parsed, dict) or not ("attributes" in parsed or "section" in parsed):
        return text

    out: list[str] = []
    accno = str(parsed.get("accno") or "").strip()
    if accno:
        out.append(f"accno={accno}")
    _flatten_section(parsed, out)
    return "\n".join(out).strip()


# The two conventions a deposit states its licence in. BioStudies writes it as an
# ATTRIBUTE — a node that names the field and carries the value beside it — while
# RO-Crate, CodeMeta, Frictionless and DataCite write it as a FIELD, the key
# itself. Both are the depositor naming a licence. Neither is prose.
_LICENCE_NAMES = {"license", "licence"}
_LICENCE_KEYS = {"license", "licence", "licenses", "licences", "rightsuri"}
# Where an object-valued licence keeps the thing itself, best first: an IRI is
# machine-actionable, a label is only a name.
_LICENCE_VALUE_KEYS = ("@id", "url", "path", "identifier", "name", "value")


def _is_iri(value: str) -> bool:
    return value.startswith(("http://", "https://"))


def _licence_candidates(value: Any) -> list[str]:
    """Every licence a value states, in document order.

    A licence is a string, an object keeping it under one of
    :data:`_LICENCE_VALUE_KEYS`, or a list of either — the shapes RO-Crate
    (``{"@id": …}``), CodeMeta, Frictionless (``{"path": …}``) and DataCite use
    between them.
    """
    if isinstance(value, str):
        return [value.strip()] if value.strip() else []
    if isinstance(value, list):
        return [found for item in value for found in _licence_candidates(item)]
    if isinstance(value, dict):
        for key in _LICENCE_VALUE_KEYS:
            if found := str(value.get(key) or "").strip():
                return [found]
    return []


# The formal, machine-readable way to declare a licence in any text at all. It
# names the field, so reading it is no more a guess than reading a JSON key.
_SPDX_DECLARATION = re.compile(r"SPDX-License-Identifier:\s*([^\s*/#<>\"\']+)", re.IGNORECASE)
# Filenames whose whole content IS the licence. The name is itself the
# declaration, which is what lets a URI be read out of the text — in any other
# file that would be a URL that happens to appear in prose.
_LICENCE_FILENAMES = {"license", "licence", "copying", "copyright"}
_URI_IN_TEXT = re.compile(r"https?://[^\s\"\'<>)\]]+")
# The XML analogue of the field convention. DataCite keeps the machine-actionable
# form in an ATTRIBUTE (`<rights rightsURI="…">CC BY 4.0</rights>`) and the label
# in the element text, while Dublin Core puts the whole thing in the text.
_LICENCE_XML_TAGS = {"license", "licence", "rights", "rightsuri"}
_LICENCE_XML_ATTRS = {"rightsuri", "href", "resource", "about", "url"}


def _local_tag(name: Any) -> str:
    """An XML tag or attribute without its namespace: ``{ns}rights`` -> ``rights``."""
    return str(name).rsplit("}", 1)[-1].strip().casefold()


def _prefer_iri(found: list[str]) -> str | None:
    """The first machine-actionable value, else the first stated one."""
    return next((value for value in found if _is_iri(value)), found[0] if found else None)


def _parse_structured(text: str) -> dict[str, Any] | None:
    """*text* as a mapping, read as JSON then as YAML, or ``None``.

    A mapping is required rather than any YAML value, because YAML parses prose
    into something: every real deposit's README carries the unfilled placeholder
    ``[Default CC-BY 4.0 for data, CC0 for metadata unless specified
    otherwise]``, which is a valid flow sequence naming two licences and
    declaring neither.
    """
    try:
        parsed = json.loads(text)
    except (ValueError, TypeError):
        try:
            parsed = yaml.safe_load(text)
        except Exception:  # noqa: BLE001 — anything unparseable is simply not a declaration
            return None
    return parsed if isinstance(parsed, dict) else None


def _licence_from_structured(text: str) -> str | None:
    """The licence a JSON/YAML mapping declares, in either convention."""
    parsed = _parse_structured(text)
    if parsed is None:
        return None

    found: list[str] = []

    def _walk(node: Any) -> None:
        if isinstance(node, dict):
            # The attribute convention: this node NAMES the licence field, and
            # a qualifier beside it may carry the canonical URL.
            if str(node.get("name") or "").strip().casefold() in _LICENCE_NAMES:
                for qualifier in node.get("valqual") or []:
                    url = str((qualifier or {}).get("value") or "").strip()
                    if _is_iri(url):
                        found.append(url)
                        return
                if value := str(node.get("value") or "").strip():
                    found.append(value)
                return
            for key, child in node.items():
                if str(key).strip().casefold() in _LICENCE_KEYS:
                    found.extend(_licence_candidates(child))
                else:
                    _walk(child)
        elif isinstance(node, list):
            for child in node:
                _walk(child)

    _walk(parsed)
    return _prefer_iri(found)


def _licence_from_xml(text: str) -> str | None:
    """The licence an XML record declares — DataCite, Dublin Core, METS.

    Parsed through ``defusedxml``: a deposit is untrusted input, and stdlib
    ElementTree expands a billion-laughs entity out of one crafted file. A
    refused or malformed document is simply not a declaration.
    """
    if "<" not in text:
        return None
    from defusedxml.ElementTree import fromstring

    try:
        root = fromstring(text)
    except Exception:  # noqa: BLE001 — unparseable or refused is not a declaration
        return None

    found: list[str] = []
    for element in root.iter():
        if _local_tag(element.tag) not in _LICENCE_XML_TAGS:
            continue
        for name, value in (element.attrib or {}).items():
            if _local_tag(name) in _LICENCE_XML_ATTRS and str(value).strip():
                found.append(str(value).strip())
        if label := (element.text or "").strip():
            found.append(label)
    return _prefer_iri(found)


_DOI_NAMES = {"doi"}
_DOI_KEYS = {"doi", "dois", "doi_url", "identifier_doi"}
# A DOI is `10.<registrant>/<suffix>`; the registrant is numeric and at least four
# digits. Deliberately narrow: this must not match a version string or a date.
_DOI_PATTERN = re.compile(r"\b(10\.\d{4,9}/[-._;()/:a-z0-9A-Z]+)")


def extract_deposit_doi(text: str) -> str | None:
    """The DOI a structured metadata document declares (#682).

    The counterpart to :func:`extract_deposit_licence`, reading the sibling attribute
    in the very same list. A BioStudies descriptor states it as

    .. code-block:: json

        {"name": "DOI", "value": "10.6019/S-VHPS22"}

    while an RO-Crate, CodeMeta or DataCite record states it as a field. Reading it is
    not guessing — the depositor named it — and it is the one thing in these deposits
    that identifies them once they leave the repository they came from: an accession
    like ``S-VHPS22`` is unique inside BioStudies and ambiguous outside it.

    **It does not infer a repository.** Nothing here maps an accession pattern onto
    ebi.ac.uk or anywhere else; a detector of that kind was deliberately removed from
    ``document_discovery`` for special-casing one repository's dialect. A DOI needs no
    such inference, which is exactly why it is the fact worth reading.

    Returned in resolvable form (``https://doi.org/…``). That is not the D5 invention
    the licence reader guards against: doi.org is the registered resolution service for
    the identifier the depositor wrote, not a claim about what the deposit contains.
    """
    parsed = _parse_structured(text)
    if parsed is None:
        return None

    found: list[str] = []

    def _walk(node: Any) -> None:
        if isinstance(node, dict):
            # The attribute convention: a node NAMING the DOI field.
            if str(node.get("name") or "").strip().casefold() in _DOI_NAMES:
                if value := str(node.get("value") or "").strip():
                    found.append(value)
                return
            for key, child in node.items():
                if str(key).strip().casefold() in _DOI_KEYS and isinstance(child, str):
                    found.append(child.strip())
                else:
                    _walk(child)
        elif isinstance(node, list):
            for child in node:
                _walk(child)

    _walk(parsed)
    for candidate in found:
        if match := _DOI_PATTERN.search(candidate):
            return "https://doi.org/" + match.group(1)
    return None


def extract_deposit_licence(text: str, *, filename: str = "") -> str | None:
    """The licence a structured metadata document declares (#535).

    Reading this is not guessing: the depositor named the field. A BioStudies
    descriptor states it as an attribute, usually qualified with a canonical
    URL —

    .. code-block:: json

        {"name": "License", "value": "CC-BY",
         "valqual": [{"name": "URL",
                      "value": "https://creativecommons.org/licenses/by/4.0/legalcode"}]}

    — while an RO-Crate, CodeMeta record, Frictionless datapackage or DataCite
    payload states it as a field. Gating on the BioStudies shape answered for
    exactly one repository's export and left every other deposit with the
    fabricated all-rights-reserved fallback, so both conventions are read.

    An IRI wins wherever it sits, being machine-actionable and the depositor's
    own choice. Without one the declared value is returned **verbatim**:
    "CC-BY" does not say which version, and mapping it onto a 4.0 URI would
    state something the depositor did not (D5).

    Two conventions live outside a metadata record and are read as well.
    ``SPDX-License-Identifier:`` is a formal declaration that can sit in any
    text, and a file *named* ``LICENSE`` / ``COPYING`` declares by its name that
    its whole content is the licence — which is what lets a URI be read out of
    it, where in any other file that would be a URL appearing in prose. Pass
    *filename* to enable that reading; without it the strict rules apply.

    Legal prose is still never mined. A ``LICENSE`` holding only the text of a
    licence names no identifier, and reading "Creative Commons Attribution 4.0
    International Public License" off its first line would invent a
    machine-actionable claim out of a heading.

    Applied blindly to any scanned file, so anything carrying no licence — or
    naming one only in prose — yields ``None`` rather than raising.
    """
    for read in (_licence_from_structured, _licence_from_xml):
        if found := read(text):
            return found
    if spdx := _SPDX_DECLARATION.search(text):
        return spdx.group(1).strip()
    if Path(filename).stem.casefold() in _LICENCE_FILENAMES:
        if uri := _URI_IN_TEXT.search(text):
            return uri.group(0)
    return None


def read_file(
    path: str,
    *,
    max_lines: int = 100,
    max_bytes: int = _MAX_BYTES,
    compact: bool = False,
) -> str | None:
    """Read a file, dispatching to the right reader based on its extension.

    Supported formats:

    - ``.txt``, ``.csv``, ``.tsv``, ``.json``, ``.yml``, ``.yaml``,
      ``.xml``, ``.md``, ``.log``, ``.ini``, ``.cfg``, ``.toml``,
      ``.py``, ``.r``, ``.sh`` — plain text, read as UTF-8
    - ``.xlsx`` / ``.xlsm`` / legacy ``.xls`` — Excel via :func:`read_excel`
    - ``.docx`` — Word via :func:`read_docx`
    - ``.pdf`` — via :func:`~builder.tools.scanner.extract_pdf_text`

    Plain-text and JSON files are returned **in full** up to a generous byte
    budget (:data:`_TEXT_BUDGET_BYTES`, 64 KiB); a file that genuinely exceeds it
    comes back with an explicit truncation marker so a weak model does not loop
    "read the rest" (Issue #240). A *directory* path returns a clear, actionable
    message (browse the inventory, then read a specific file) instead of *None*.

    Unsupported or unreadable files return *None*.

    Args:
        path: Path to the file.
        max_lines: Max rows to return from structured formats (``.xlsx``); text
            and JSON are governed by the byte budget, not a line cap.
        max_bytes: Hard safety cap in bytes (default 100 MB); files larger than
            this are skipped (returns *None*).
        compact: Apply the shared boilerplate compactors (#378) —
            :func:`compact_grid_text` for ``.xlsx`` grids and
            :func:`compact_attribute_json` for BioStudies ``.json`` descriptors.
            Defaults to **False** so no existing caller's output changes; the
            deterministic pipeline opts in for its bounded context slice.

    Returns:
        File content as a string, a directory-guidance message, or *None*.
    """
    file_path = Path(path)
    if file_path.is_dir():
        return _directory_message(path)
    if not file_path.is_file():
        return None

    suffix = file_path.suffix.lower()

    if suffix in (".xlsx", ".xlsm", ".xls"):
        return read_excel(path, max_rows=max_lines, max_bytes=max_bytes, compact=compact)

    if suffix == ".docx":
        return read_docx(path, max_bytes=max_bytes)

    if suffix == ".pdf":
        from builder.tools.scanner import extract_pdf_text

        return extract_pdf_text(path)

    # Everything else is read as plain text (full content up to the byte budget).
    text = _read_text_file(path, max_bytes=max_bytes)
    if compact and text and suffix == ".json":
        return compact_attribute_json(text)
    return text


# ---------------------------------------------------------------------------
# Explicit ToolRegistry registration
# ---------------------------------------------------------------------------

from builder.tools.registry import TOOL_REGISTRY  # noqa: E402

TOOL_REGISTRY.register(
    "read_excel",
    read_excel,
    description="Read an Excel .xlsx file and return its content as pipe-delimited text",
)
TOOL_REGISTRY.register(
    "read_docx",
    read_docx,
    description="Read a Word .docx file and return its text content",
)
# A worksheet whose name marks it as the run's plan. Observed as `layout 27-03`,
# `layout 01-03` and `Layout` across this lab's deposits — the date suffix varies
# per run, so the match is a prefix-free substring on the case-folded name.
_LAYOUT_SHEET = "layout"


def read_layout_conditions(path: Any) -> dict[str, str]:
    """Conditions a run's ``layout`` sheet states for the whole run (#697).

    Every experiment workbook in this lab's deposits opens its layout sheet with
    a block of label/value pairs — incubation volume, buffer, substrate, dose,
    duration — which are exactly the columns an empty condition table is missing.
    They describe the run as a whole rather than a well, so they belong on the
    Exposure as parameters rather than as table rows.

    **Bounded to the leading block.** Below it sits the design matrix, whose
    cells are adjacent pairs too, so a scan of the whole sheet turns ``D | D``
    into a condition named "D". The block is the contiguous run of pairs from the
    top, ending at the first blank row after it starts; a leading blank row is
    skipped, because every real sheet has one. That bound is the sheet's own
    structure rather than a list of labels to expect, so a run stating something
    new is read rather than filtered.

    Only the values are taken. Nothing here parses ``1 nM T3 or T4`` into a
    quantity or resolves ``Xn`` to a compound: those are the depositor's words,
    carried verbatim, and inventing structure for them would be a guess.

    Never raises — a deposit holds workbooks no reader can open, and one of them
    must not stop a build.

    Args:
        path: The workbook to read.

    Returns:
        ``{label: value}`` in sheet order, or ``{}`` when the workbook has no
        layout sheet, no block, or cannot be opened.
    """
    from pathlib import Path as _Path

    try:
        import openpyxl
    except ImportError:  # pragma: no cover — openpyxl is a hard dependency
        return {}
    _silence_openpyxl_extension_warnings()
    try:
        workbook = openpyxl.load_workbook(_Path(path), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — an unreadable workbook is not an error
        logger.debug("No layout conditions from %s: %s", path, exc)
        return {}
    try:
        sheets = [n for n in workbook.sheetnames if _LAYOUT_SHEET in n.casefold()]
        if not sheets:
            return {}
        return _leading_pairs(workbook[sheets[0]])
    except Exception as exc:  # noqa: BLE001 — same reason
        logger.debug("No layout conditions from %s: %s", path, exc)
        return {}
    finally:
        workbook.close()


def _leading_pairs(worksheet: Any) -> dict[str, str]:
    """The contiguous label/value block at the top of *worksheet*."""
    found: dict[str, str] = {}
    started = False
    for row in worksheet.iter_rows(values_only=True):
        pair = None
        cells = list(row[:4])
        for left, right in zip(cells, cells[1:]):
            if not isinstance(left, str) or not left.strip():
                continue
            if right is None or not str(right).strip():
                continue
            pair = (left.strip(), str(right).strip())
            break
        if pair is None:
            # A blank row before anything has been read is the sheet's own
            # leading gap; one after the block has started is its end.
            if started:
                break
            continue
        started = True
        found.setdefault(pair[0], pair[1])
    return found


def shared_layout_conditions(paths: Any) -> dict[str, str]:
    """What EVERY run in *paths* states identically (#697).

    An assay holds several experiment workbooks and their blocks disagree: one
    run of the metabolism assay used ``H4 + SKNAS`` and another ``MO3.13``, one
    uptake run inhibited with ``Xn`` and another with ``none``. Only what every
    run agrees on is a property of the assay; the rest is per-run and belongs
    with the work that gives each run its own exposed samples (#654).

    Agreement is exact on the stated text. ``24hour`` and ``24 hours`` mean the
    same thing and do not say so — normalising them would be a guess about units
    and format, so they are dropped rather than reconciled. A label only some
    runs state is dropped for the same reason: silence is not agreement.

    Args:
        paths: Workbooks belonging to one assay.

    Returns:
        ``{label: value}`` every run stated identically, or ``{}``.
    """
    blocks = [read_layout_conditions(path) for path in paths]
    blocks = [b for b in blocks if b]
    if not blocks:
        return {}
    first, rest = blocks[0], blocks[1:]
    return {
        label: value
        for label, value in first.items()
        if all(other.get(label) == value for other in rest)
    }


TOOL_REGISTRY.register(
    "read_file",
    read_file,
    description="Read any supported file format by extension (txt, csv, json, xlsx, docx, md, pdf). Text/JSON come back in full up to 64 KiB; a larger file is truncated with an explicit marker, and a directory returns guidance to use list_scanned_files",  # noqa: E501
)
