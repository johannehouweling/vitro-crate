#!/usr/bin/env python3
"""Generate ``fair/dsm_indicators.yaml`` from the vendored FAIRplus DSM assessment sheet.

The FAIRplus **Dataset** Maturity Model (DSM) indicator *definitions* — id, level,
category, verbatim text, granularity, and the model's own cross-references to RDA and
FAIRsFAIR indicators — are sourced from the canonical assessment workbook
(``fair/fairplus_dsm_v1.2.xlsx``, FAIRplus/Data-Maturity, model text CC BY 4.0) rather than
hand-authored, so they cannot silently drift. This mirrors how
``scripts/gen_fair_indicators.py`` sources the RDA indicators from ``fair/rda_fdmm.xlsx``.

The *local* decision — which indicators this tool can assess intrinsically from a single
RO-Crate, and with which check function — lives in :data:`LOCAL_SCOPE` below; that is
inherently repo-specific and stays here.

**All 83 published indicators are carried**, including the ones we cannot assess. An
indicator we do not assess is reported ``na`` (not assessable), never failed — the
published model is reproduced in full and our coverage of it is visible rather than
implied. Three groups are ``na`` by construction:

* **Hosting Env. Capabilities** — the model's own ``Granularity`` column marks these
  *Storage / Retrieval / Search Capability*, i.e. properties of the environment serving
  the dataset, not of the dataset. A crate on disk cannot evidence them.
* **Level 0** — these are *negative* statements describing the pre-FAIRification state
  ("Dataset(s) are NOT Identifiable via Unique Identifiers"). They are a starting
  condition, not a target to satisfy.
* Content/Representation indicators needing file-content or external context we do not
  inspect.

Regenerate with::

    uv run python scripts/gen_dsm_indicators.py

``tests/test_dsm_indicators_source.py`` asserts the committed file equals this
generator's output and that every indicator matches its row in the workbook.
"""

from __future__ import annotations

import pathlib
import re
from typing import Any

import openpyxl
import yaml

REPO = pathlib.Path(__file__).resolve().parents[1]
DSM_XLSX = REPO / "fair" / "fairplus_dsm_v1.2.xlsx"
OUT = REPO / "fair" / "dsm_indicators.yaml"

# The workbook's authoritative indicator listing. Columns: Level(0), Category(1),
# ID(2), v1 Identifier(3), Indicator Description(4), REF RDA indicator(s)(5),
# REF FAIRsFAIR indicator(s)(6), Granurality(7)  [sic — the sheet's own spelling].
_SHEET = "MASTER (Levels View)"

# The scoring sheet. Column H is the pre-FAIRification assessment, I the post; J and K
# are the "Validation" columns that turn those answers into the published % grid. J is
# ENTIRELY formulas (`=H{row}`, or `=IF(J{src}=1,1,H{row})` where a higher rung
# satisfies a lower one), so a blank H evaluates to numeric 0 and COUNT() counts it:
# this instrument has no "not assessed" state, and every denominator equals its member
# count. Columns N/O/P hold the % grid — level, category, formula.
_ASSESSMENT_SHEET = "FAIR-DSM Assessment Sheet v1.2"

# The grid's own category labels -> the codes the indicator ids use. "ant" is the
# sheet's own typo and is matched verbatim; TOTAL is a fourth column, not a category.
_GRID_CATEGORY_CODE: dict[str, str] = {
    "Content ant context": "C",
    "Representation and format": "R",
    "Hosting environment capabilities": "H",
    "Total": "TOTAL",
}

# The workbook's category names -> the single-letter codes the indicator ids use.
_CATEGORY_CODE: dict[str, str] = {
    "Content & Context": "C",
    "Representation & Format": "R",
    "Hosting Env. Capabilities": "H",
}

# Level names, verbatim from the workbook's "Levels definition" sheet ("State of
# Data"). These name the state of the DATA, not an organisation's capability — the
# FAIR Cookbook is explicit that the model defines "maturity levels which can used to
# describe a dataset".
_LEVELS: dict[int, str] = {
    0: "Single-use data",
    1: "Identifiable Data",
    2: "Described Data",
    3: "Standardised Data",
    4: "Semantically Typed Data",
    5: "Managed Data Assets",
}

# The crate-intrinsic subset this tool assesses: published indicator id -> check
# function name (see builder/tools/fair_assessment.py DSM_CHECKS). Every id here is
# verified against the workbook at generation time, so a typo or an indicator the
# model drops fails loudly rather than silently scoring nothing.
#
# DSM-3-C1 ("study-level metadata is reported in compliance with relevant Minimum
# Information Reporting Guidelines") is NOT here, and that is a measurement rather than
# an oversight. Compliance means a guideline's required items are all reported, and the
# MIT checklist this tool scores carries no such threshold: every parameter a
# document flags is flagged equally, and 7 to 24 parameters per document have no crate
# slot at all, so 100% is unreachable by construction — the best-reported crate on the
# corpus tops out at 29 of 48 on OHT 201 and 4 of 7 on Nature. Every predicate that
# separates a real crate from an empty one is a shape-of-coverage claim ("no document is
# at zero"), which is not the published question. So the crate does not answer this one;
# the depositor does, like every other indicator a crate cannot evidence (#705).
#
# `scope: full`    — directly decidable from the assembled @graph.
# `scope: partial` — a presence/structure heuristic; true semantic compliance would
#                    need content or external checks we do not perform.
# Anything absent here is `scope: na`.
LOCAL_SCOPE: dict[str, tuple[str, str]] = {
    # ---- Level 1 — identifiable, access info, machine-readable descriptor --------
    "DSM-1-C0": ("full", "unique_id"),
    "DSM-1-C1": ("full", "study_summary"),
    "DSM-1-C2": ("full", "dataset_metadata"),
    "DSM-1-C3": ("full", "access_info"),
    "DSM-1-R0": ("full", "has_descriptor"),
    "DSM-1-R1": ("full", "context_fields"),
    "DSM-1-R2": ("full", "dataset_hierarchy"),
    "DSM-1-R3": ("full", "general_schema"),
    "DSM-1-R4": ("full", "descriptor_machine_readable"),
    "DSM-1-R5": ("full", "data_machine_readable"),
    # ---- Level 2 — field/value metadata, local domain + dataset model ------------
    "DSM-2-C1": ("partial", "domain_model"),
    "DSM-2-C2": ("partial", "tidy_dataset"),
    "DSM-2-C3": ("full", "reference_fields"),
    "DSM-2-C4": ("partial", "local_data_dictionary"),
    "DSM-2-C5": ("full", "cross_dataset_refs"),
    "DSM-2-C6": ("full", "field_level_metadata"),
    "DSM-2-C7": ("full", "value_level_metadata"),
    "DSM-2-R1": ("partial", "domain_model"),
    "DSM-2-R2": ("full", "local_dataset_model"),
    "DSM-2-R3": ("full", "generic_model"),
    "DSM-2-R4": ("partial", "model_documentation_human"),
    "DSM-2-R5": ("full", "data_structured"),
    # ---- Level 3 — community standards, controlled terms, standard licence -------
    # DSM-3-C1 is deliberately absent: see the note above LOCAL_SCOPE.
    "DSM-3-C2": ("partial", "community_domain_model"),
    "DSM-3-C3": ("full", "standard_field_names"),
    "DSM-3-C4": ("partial", "controlled_values"),
    "DSM-3-C5": ("full", "resolvable_terms"),
    "DSM-3-C6": ("partial", "standard_field_metadata"),
    "DSM-3-C7": ("full", "standard_license"),
    "DSM-3-R1": ("partial", "community_domain_model"),
    "DSM-3-R2": ("full", "local_dataset_model"),
    "DSM-3-R3": ("full", "domain_standard"),
    "DSM-3-R4": ("full", "community_domain_model"),
    "DSM-3-R5": ("full", "non_proprietary_format"),
    # ---- Level 4 — semantic model, linked data, machine-readable licence ---------
    "DSM-4-C1": ("partial", "semantic_study_design"),
    "DSM-4-C2": ("partial", "common_data_elements"),
    "DSM-4-C3": ("full", "common_data_elements"),
    "DSM-4-C4": ("partial", "standard_identifiers"),
    "DSM-4-C5": ("partial", "cde_relationships"),
    "DSM-4-R1": ("full", "semantic_contextual_metadata"),
    "DSM-4-R2": ("partial", "linked_data"),
    "DSM-4-R3": ("partial", "semantic_model"),
    "DSM-4-R4": ("partial", "machine_interpretable"),
    "DSM-4-R5": ("partial", "machine_interpretable_graph"),
    "DSM-4-R6": ("full", "license_machine"),
}

# The action that clears each assessable indicator, and the one consequence a reuser
# feels while it is open. Repo-authored like LOCAL_SCOPE — the workbook states the
# question, never the fix — and written against WHAT THE CHECK MEASURES, so a depositor
# who does this sees the indicator turn.
#
# `do` is an imperative naming the property or entity; `why` is a consequence for a
# reuser, never a restatement ("so the indicator passes" says nothing). Both are drawn
# in the report exactly where a validator finding draws its instruction and its reason,
# so the two instruments give one kind of advice.
REMEDIES: dict[str, tuple[str, str]] = {
    # ---- Level 1 ----------------------------------------------------------------
    "DSM-1-C0": (
        "Mint a DOI for the deposit and record it on the root as `identifier`.",
        "a reader cannot cite the deposit or fetch it again once it leaves this crate",
    ),
    "DSM-1-C1": (
        "Write the root a `description` that says more than its name, and describe every "
        "Study Dataset.",
        "a reuser meets a named study with nothing said about what was done",
    ),
    "DSM-1-C2": (
        "Give every Dataset the root gathers, the root included, its own name and description.",
        "a reuser opening a Dataset finds no statement of what it holds",
    ),
    "DSM-1-C3": (
        "Record the repository accession for the deposit, so the descriptor says where the "
        "data can be got.",
        "a reuser holds the data with no handle for going back to the original deposit",
    ),
    "DSM-1-R0": (
        "Name the deposit — set the crate's title, which is what the descriptor identifies "
        "it by.",
        "a reuser cannot tell one deposit's descriptor from another's",
    ),
    "DSM-1-R1": (
        "Draft the study's subjects into the crate — the cell lines, the test chemicals, the "
        "protocols.",
        "a reuser opens a container that records nothing but its own packaging",
    ),
    "DSM-1-R2": (
        "File each deposited file under an Assay or Study Dataset below the root, not on the "
        "root itself.",
        "a reuser cannot tell which study or assay a file belongs to",
    ),
    "DSM-1-R3": (
        "Fill `name`, `description`, `datePublished` and `license` on the Root Data Entity.",
        "a generic schema.org reader finds no title, terms or date for the deposit",
    ),
    "DSM-1-R4": (
        "Declare the RO-Crate profile IRI as `conformsTo` on `ro-crate-metadata.json`.",
        "a machine reading the file cannot tell which specification it follows",
    ),
    "DSM-1-R5": (
        "Deposit the tables as CSV, TSV or XLSX and declare each file's `encodingFormat`.",
        "software cannot read the numbers out of a PDF or an undeclared blob",
    ),
    # ---- Level 2 ----------------------------------------------------------------
    "DSM-2-C1": (
        "Write the study design, and how its datasets relate, into the root's `description`.",
        "a reuser cannot tell what the study did or how its files fit together",
    ),
    "DSM-2-C2": (
        "Declare a `datatype` on every `csvw:Column`, and describe at least two columns in all.",
        "an untyped column is an undifferentiated string to every tool that reads it",
    ),
    "DSM-2-C3": (
        "Give at least one column a `valueUrl` pointing at the Sample or MolecularEntity its "
        "values name.",
        "a reuser cannot join the table's rows to the samples and compounds measured",
    ),
    "DSM-2-C4": (
        "Give at least one column both a `datatype` and a `valueUrl` or `propertyUrl` naming "
        "its vocabulary.",
        "a reuser has to guess what the codes in a column stand for",
    ),
    "DSM-2-C5": (
        "Hang the depositor's own files directly under at least two Datasets the root relates to.",
        "a reuser cannot see which datasets the deposit holds, or how they join",
    ),
    "DSM-2-C6": (
        "Point every deposited table at a `tableSchema` listing one column per field, each with "
        "`titles` and a `datatype`.",
        "the deposit's tables arrive as bare headers a reader has to decode by hand",
    ),
    "DSM-2-C7": (
        "Record at least two property values on at least one entity in the crate.",
        "an entity carrying a single value states nothing a reuser can interpret",
    ),
    "DSM-2-R1": (
        "Fill the root's `description` with the project's contextual model: design, arms, "
        "endpoints.",
        "the context that makes the numbers mean anything is nowhere in the crate",
    ),
    "DSM-2-R2": (
        "Declare a `tableSchema` on at least one data table in the crate.",
        "a reuser opening the CSV has no column list to read the numbers against",
    ),
    "DSM-2-R3": (
        "Record at least one entity — a sample, a compound, an assay — before exporting.",
        "an empty descriptor models no data at all",
    ),
    "DSM-2-R4": (
        "Write a `description` on at least one LabProtocol, not just its name.",
        "a reuser gets a protocol title and no account of what was done",
    ),
    "DSM-2-R5": (
        "Give every deposited file a specific `encodingFormat` rather than none or "
        "`application/octet-stream`, and leave no Dataset with an empty `hasPart`.",
        "a machine cannot tell how to open the file, and a promised dataset turns out empty",
    ),
    # ---- Level 3 ----------------------------------------------------------------
    "DSM-3-C2": (
        "Declare the community profile the datasets follow in the root's `conformsTo`.",
        "a reader cannot tell which community guideline the datasets were reported under",
    ),
    "DSM-3-C3": (
        "Give each column a `propertyUrl` that is an absolute ontology IRI, not a local name.",
        "a reuser cannot tell what a column heading means across datasets",
    ),
    "DSM-3-C4": (
        "Bind every MolecularEntity and cell-line Sample to a registry IRI via `@id`, "
        "`identifier`, `sameAs` or `url`.",
        "a reuser cannot tell which cell line or compound was actually used",
    ),
    "DSM-3-C5": (
        "Give a reported term — cell line, compound, AOP — a `url` or `sameAs` holding the full "
        "registry link, not a bare accession.",
        "a reuser cannot look up the term behind a bare label",
    ),
    "DSM-3-C6": (
        "Write the per-well design rows into the condition table, so the CSV carries records and "
        "not only a header.",
        "a schema over an empty table tells a reuser nothing about the data",
    ),
    "DSM-3-C7": (
        "Set the root's `license` to a standard licence URL — a Creative Commons, SPDX or OSI IRI.",
        "nobody may legally reuse the data without one",
    ),
    "DSM-3-R1": (
        "Declare in the root's `conformsTo` the community standard the contextual metadata "
        "follows.",
        "a reuser cannot tell which standard the context was modelled against",
    ),
    "DSM-3-R2": (
        "Attach a `csvw:tableSchema` to the crate's tabular data files.",
        "a reuser cannot parse the columns without a declared schema",
    ),
    "DSM-3-R3": (
        "Declare a domain profile IRI in the root's `conformsTo` alongside the RO-Crate one.",
        "a reader sees packaging conventions but no domain standard to read against",
    ),
    "DSM-3-R4": (
        "Publish the adopted model as a resolvable profile IRI and declare it in `conformsTo`.",
        "a machine has no address to fetch the model the crate claims to follow",
    ),
    "DSM-3-R5": (
        "Deposit the data as CSV, TSV, JSON, XML, HDF5 or NetCDF and state that `encodingFormat`.",
        "a reuser needs licensed software to open .xlsx or .prism data",
    ),
    # ---- Level 4 ----------------------------------------------------------------
    "DSM-4-C1": (
        "Wire each LabProcess to its object and its result, not only to input and output.",
        "a reuser cannot see which material a step consumed or what it produced",
    ),
    "DSM-4-C2": (
        "Give most of the crate's columns a `propertyUrl` naming an ontology term, and at "
        "least two.",
        "an unmapped column leaves a reuser guessing what the field records",
    ),
    "DSM-4-C3": (
        "Map the majority of columns to ontology terms via `propertyUrl`, not only the ones you "
        "consider key.",
        "a column no shared term names cannot be joined to anyone else's table",
    ),
    "DSM-4-C4": (
        "Carry the Cellosaurus or DTXSID identifier onto every node the tables point at, the "
        "minted output Sample of each culture step included.",
        "a reuser cannot tell which cells the measured wells actually held",
    ),
    "DSM-4-C5": (
        "Declare the study's controlled terms as DefinedTerm entities and reference them from the "
        "entities that use them.",
        "loose terms leave a reuser unable to tell which concept a value names",
    ),
    "DSM-4-R1": (
        "Give every PropertyValue a `propertyID` that is a resolvable ontology IRI.",
        "a bare label tells a machine nothing about which property was recorded",
    ),
    "DSM-4-R2": (
        "Attach a `tableSchema` to every deposited CSV and set `propertyUrl` on each of its "
        "columns.",
        "an undescribed table cannot be queried or joined alongside other data",
    ),
    "DSM-4-R3": (
        "Describe at least one study entity — a sample, a protocol, an assay process — so the "
        "crate holds more than a file listing.",
        "a crate of packaging metadata alone tells a reuser nothing about the study",
    ),
    "DSM-4-R4": (
        "Declare a `datatype` on every column, and give each one either a `valueUrl` or a "
        "non-string datatype such as `xsd:double`.",
        "a bare string column tells a machine nothing about what its values denote",
    ),
    "DSM-4-R5": (
        "Bind each column's `propertyUrl` to an external ontology IRI.",
        "a reuser's software cannot tell what any column measures",
    ),
    "DSM-4-R6": (
        "Set the root's `license` to the licence URL itself, not a label like \"CC-BY\".",
        "a bare label does not say which version governs reuse",
    ),
}


SOURCE: dict[str, Any] = {
    "name": "FAIRplus Dataset Maturity Model (DSM)",
    "url": "https://fairplus.github.io/Data-Maturity/",
    "repository": "https://github.com/FAIRplus/Data-Maturity",
    "distribution": (
        "docs/assessment/FAIR-DSM-Assessment-Sheet-v1.2.xlsx "
        "(sheet 'MASTER (Levels View)') — vendored as fair/fairplus_dsm_v1.2.xlsx"
    ),
    "version": "1.2",
    # The MODEL TEXT is CC BY 4.0 and requires attribution. The repository's
    # LICENSE.txt is only the MIT licence of its Just-the-Docs Jekyll theme
    # ((c) Patrick Marsceill) and does NOT license the model — a trap worth naming,
    # because taking the repo's headline licence at face value would under-attribute
    # a work this paper cites.
    "license": "CC-BY-4.0",
    "specification": {
        "name": "FAIRplus D2.6 FAIR Data Set Maturity model",
        "doi": "10.5281/zenodo.7464523",
        "url": "https://doi.org/10.5281/zenodo.7464523",
        "year": 2022,
        "authors": (
            "Emam I, Rocca-Serra P, Sansone S-A, Portell-Silva L, Gadiya Y, "
            "Welter D, Juty N, Abbassi-Daloii T"
        ),
    },
    "peer_reviewed": {
        "citation": (
            "Welter D, Juty N, Rocca-Serra P, et al. FAIR in action - a flexible "
            "framework to guide FAIRification. Sci Data 10:291 (2023)"
        ),
        "doi": "10.1038/s41597-023-02167-2",
    },
    "assessment_tool": "https://fairdsm.biospeak.solutions/assess",
    # The DSM is built on the RDA FAIR Data Maturity Model: the workbook ships an
    # "RDA indicators" sheet and 27 of the 83 indicators carry an explicit
    # `REF RDA indicator(s)` cross-reference, surfaced per indicator as `rda_ref`.
    "derived_from": {
        "name": "RDA FAIR Data Maturity Model",
        "doi": "10.15497/rda00050",
        "url": "https://doi.org/10.15497/rda00050",
    },
    "retrieved": "2026-08-21",
    "note": (
        "The model assesses a DATASET, not an organisation: the FAIR Cookbook states it "
        "defines 'maturity levels which can used to describe a dataset'. Level names are "
        "the workbook's own 'State of Data'. Indicators this tool cannot assess from a "
        "crate alone are reported na, never failed."
    ),
}

_HEADER = """\
# GENERATED FILE — do not edit by hand.
#
# Regenerate with:  uv run python scripts/gen_dsm_indicators.py
#
# All 83 published FAIRplus Dataset Maturity Model (DSM) indicators, sourced verbatim
# from the vendored assessment workbook (fair/fairplus_dsm_v1.2.xlsx; model text
# CC BY 4.0, FAIRplus D2.6, doi:10.5281/zenodo.7464523).
# `text` is the model's own wording; `granularity`, `rda_ref` and `fairsfair_ref` are
# the model's own columns. The local scope/check mapping — which indicators this tool
# assesses from one RO-Crate, and with which check — lives in
# scripts/gen_dsm_indicators.py. Hosting-environment indicators and the Level 0
# pre-FAIRification states are reported na (not assessable), never failed.
"""


def _load_workbook_rows() -> list[dict[str, Any]]:
    """Every DSM indicator in the workbook, in sheet order.

    The Level column is only filled on the first row of each level block, so it is
    carried down. Rows without a ``DSM-`` id are spacers and are skipped.
    """
    wb = openpyxl.load_workbook(DSM_XLSX, read_only=True, data_only=True)
    ws = wb[_SHEET]
    rows: list[dict[str, Any]] = []
    level: int | None = None
    for row in list(ws.iter_rows(values_only=True))[1:]:
        if row[0] not in (None, ""):
            try:
                level = int(float(row[0]))
            except (TypeError, ValueError):  # a stray note in the Level column
                pass
        ident = str(row[2]).strip() if row[2] else ""
        if not ident.startswith("DSM-"):
            continue
        category = str(row[1]).strip() if row[1] else ""
        if category not in _CATEGORY_CODE:
            raise KeyError(f"{ident}: unknown DSM category {category!r}")
        if level is None:
            raise ValueError(f"{ident}: no level in scope — the sheet layout changed")
        rows.append(
            {
                "id": ident,
                "level": level,
                "category": _CATEGORY_CODE[category],
                "text": str(row[4]).strip() if row[4] else "",
                "rda_ref": str(row[5]).strip() if row[5] else "",
                "fairsfair_ref": str(row[6]).strip() if row[6] else "",
                "granularity": str(row[7]).strip() if row[7] else "",
            }
        )
    return rows


# The questionnaire sheet. Each row is one ANSWER OPTION; rows sharing a QUESTION are
# that question's options, ordered as an increasing ladder (one statement per maturity
# level). Columns: CATEGORY(1), SUBCATEGORY(2), QUESTION_NUM(5), QUESTION(6),
# INDICATORID(8), INDICATOR(9), LEVEL(10), ISMULTI(11), HAS_NOA(12).
_QUESTION_SHEET = "Assessment Tool Data v1.2 "


def _assessment_rows(ws: Any) -> dict[int, str]:
    """Sheet row number -> indicator id, for the rows the scoring formulas reference."""
    rows: dict[int, str] = {}
    for row in range(2, ws.max_row + 1):
        ident = ws.cell(row=row, column=5).value  # E: "Indicator Identifier v1"
        ident = str(ident).strip() if ident else ""
        if ident.startswith("DSM-"):
            rows[row] = ident
    return rows


def _promotion_rules(ws: Any, rows: dict[int, str]) -> list[dict[str, str]]:
    """The sheet's own ladder, read off the J column.

    ``=IF(J5=1,1,H4)`` says: if the rung recorded on row 5 is met, row 4 counts as met
    too. The model's statements nest — "standardised to a *community* model" implies
    "standardised to a *local* model" — and the sheet resolves that by PROMOTING the
    lower rung, never by demoting the higher one. Anything else in the column is a
    layout change we must not paper over, so it raises.
    """
    plain = re.compile(r"^=H(\d+)$")
    promote = re.compile(r"^=IF\(J(\d+)=1,1,H(\d+)\)$")
    out: list[dict[str, str]] = []
    for row, ident in rows.items():
        formula = str(ws.cell(row=row, column=10).value or "").replace(" ", "")
        mirror = str(ws.cell(row=row, column=11).value or "").replace(" ", "")
        if mirror != formula.replace("J", "K").replace("H", "I"):
            raise ValueError(
                f"row {row} ({ident}): the pre column J and the post column K disagree "
                f"({formula!r} vs {mirror!r}) — the workbook changed shape"
            )
        if direct := plain.match(formula):
            if int(direct.group(1)) != row:
                raise ValueError(f"row {row} ({ident}): {formula!r} reads another row")
            continue
        match = promote.match(formula)
        if not match:
            raise ValueError(f"row {row} ({ident}): unrecognised validation formula {formula!r}")
        source, own = int(match.group(1)), int(match.group(2))
        if own != row:
            raise ValueError(f"row {row} ({ident}): {formula!r} falls back to another row")
        out.append({"cell": f"J{row}", "then": ident, "when": rows[source]})
    return out


def _grid_cells(ws: Any, rows: dict[int, str]) -> list[dict[str, Any]]:
    """The published "% Complete" grid: level x {C, R, H, Total}, from the P column.

    Membership is the sheet's, not ours: a level's cell carries lower-level indicators
    forward (``P14`` counts DSM-1-C2 and DSM-1-C3 alongside the Level-2 rows), and it
    is a MULTISET — ``P24`` names DSM-4-H2 on two rows and divides by 3, so deduping
    would change the denominator. Level 0 counts zeros, because its statements are the
    pre-FAIRification condition stated in the negative.
    """
    counts = re.compile(r"COUNTIFS\(J(\d+),(\d)\)")
    denom_count = re.compile(r"COUNT\(([^)]*)\)")
    denom_literal = re.compile(r"/(\d+)\)\*100")
    cells: list[dict[str, Any]] = []
    level: int | None = None
    for row in range(2, ws.max_row + 1):
        label = ws.cell(row=row, column=14).value  # N: "Level 3"
        if label and str(label).strip().startswith("Level"):
            level = int(str(label).strip().split()[-1])
        category = str(ws.cell(row=row, column=15).value or "").strip()  # O
        if not category:
            continue
        if category not in _GRID_CATEGORY_CODE:
            raise KeyError(f"P{row}: unknown grid category {category!r}")
        if level is None:
            raise ValueError(f"P{row}: no level in scope — the sheet layout changed")
        cell: dict[str, Any] = {
            "cell": f"P{row}",
            "level": level,
            "category": _GRID_CATEGORY_CODE[category],
        }
        formula = ws.cell(row=row, column=16).value
        if isinstance(formula, (int, float)):
            # The sheet hardcodes P28 to 0: there is no Level-5 hosting row to count.
            cell["constant"] = float(formula)
            cell["rows"] = []
            cell["members"] = []
            cells.append(cell)
            continue
        text = str(formula).replace(" ", "").replace("++", "+")
        members = counts.findall(text)
        if not members:
            raise ValueError(f"P{row}: no COUNTIFS terms in {text!r}")
        criteria = {digit for _, digit in members}
        if len(criteria) != 1:
            raise ValueError(f"P{row}: mixed COUNTIFS criteria {criteria}")
        cell["counts"] = int(criteria.pop())
        cell["rows"] = [int(r) for r, _ in members]
        cell["members"] = [rows[int(r)] for r, _ in members]
        if match := denom_count.search(text):
            cell["denominator"] = {"kind": "count", "n": len(match.group(1).split(","))}
        elif match := denom_literal.search(text):
            cell["denominator"] = {"kind": "literal", "n": int(match.group(1))}
        else:
            raise ValueError(f"P{row}: no denominator in {text!r}")
        if len(cell["members"]) != cell["denominator"]["n"]:
            raise ValueError(
                f"P{row}: {len(cell['members'])} members but denominator "
                f"{cell['denominator']['n']} — the two arithmetic paths have diverged"
            )
        cells.append(cell)
    return cells


def _load_scoring(known: set[str]) -> dict[str, Any]:
    """The workbook's own arithmetic: the ladder, the grid, and what it leaves out.

    Carrying this means the scorer reproduces the published instrument instead of
    hand-coding an approximation of it. The cross-checks below are the parity proof:
    the transitive closure of the sheet's IF-chain must equal the ONLINE tool's
    ``SUPERSEDED_BY`` column, two independently maintained authorities agreeing.
    """
    wb = openpyxl.load_workbook(DSM_XLSX, data_only=False)
    ws = wb[_ASSESSMENT_SHEET]
    rows = _assessment_rows(ws)
    promotion = _promotion_rules(ws, rows)
    grid = _grid_cells(ws, rows)

    direct: dict[str, set[str]] = {}
    for rule in promotion:
        direct.setdefault(rule["then"], set()).add(rule["when"])

    def closure(ident: str) -> set[str]:
        seen: set[str] = set()
        stack = list(direct.get(ident, ()))
        while stack:
            node = stack.pop()
            if node not in seen:
                seen.add(node)
                stack.extend(direct.get(node, ()))
        return seen

    published = _superseded_by()
    derived = {ident: closure(ident) for ident in direct}
    if derived != published:
        raise ValueError(
            "the sheet's promotion chain and the online tool's SUPERSEDED_BY column "
            f"disagree: {derived} vs {published}"
        )

    for level in {cell["level"] for cell in grid}:
        at = {cell["category"]: set(cell["members"]) for cell in grid if cell["level"] == level}
        union = at.get("C", set()) | at.get("R", set()) | at.get("H", set())
        if union != at.get("TOTAL", set()):
            raise ValueError(f"level {level}: the Total cell is not the union of C/R/H")

    in_a_cell = {member for cell in grid for member in cell["members"]}
    off_sheet = sorted(known - in_a_cell)
    scored_off_sheet = sorted(set(LOCAL_SCOPE) - in_a_cell)
    if scored_off_sheet != ["DSM-2-R5"]:
        raise ValueError(
            "indicators we score that no published cell counts changed: "
            f"{scored_off_sheet} (expected only DSM-2-R5)"
        )
    return {
        "sheet": _ASSESSMENT_SHEET,
        "column": "J",
        "note": (
            "Column J is entirely formulas, so a blank H is a NUMERIC 0, not a blank. "
            "COUNT() counts it: this instrument has no 'not assessed' state, and every "
            "denominator equals its member count."
        ),
        "promotion": promotion,
        "grid": grid,
        "off_sheet": off_sheet,
    }


def _superseded_by() -> dict[str, set[str]]:
    """The questionnaire sheet's own statement of which rung supersedes which.

    Read with ``data_only=True``: the INDICATORID column is a cross-sheet formula, so
    the formula view yields ``='MASTER (Levels View)'!C6`` rather than the id.
    """
    wb = openpyxl.load_workbook(DSM_XLSX, read_only=True, data_only=True)
    ws = wb[_QUESTION_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]
    col = {name: i for i, name in enumerate(header) if name}
    out: dict[str, set[str]] = {}
    for row in rows[1:]:
        ident = str(row[col["INDICATORID"]]).strip() if row[col["INDICATORID"]] else ""
        value = row[col["SUPERSEDED_BY"]]
        if ident.startswith("DSM-") and value:
            out[ident] = {part.strip() for part in str(value).split(",") if part.strip()}
    return out


def build_data() -> dict[str, Any]:
    """The full ``dsm_indicators.yaml`` payload: source + levels + all indicators."""
    rows = _load_workbook_rows()
    known = {r["id"] for r in rows}
    unknown = sorted(set(LOCAL_SCOPE) - known)
    if unknown:
        raise KeyError(
            f"LOCAL_SCOPE names indicators absent from the workbook: {unknown}. "
            "Fix the mapping rather than the workbook."
        )

    # An indicator we assess is an indicator a crate can fail, so it must be able to
    # say what to do about it. Without this the report names a blocker and leaves the
    # reader to work the fix out from the model's own question, which is what the
    # validator findings stopped doing at #607.
    unremedied = sorted(set(LOCAL_SCOPE) - set(REMEDIES))
    stray = sorted(set(REMEDIES) - set(LOCAL_SCOPE))
    if unremedied or stray:
        raise KeyError(
            f"REMEDIES must name exactly the assessable indicators. Missing: {unremedied}; "
            f"naming indicators we do not assess: {stray}."
        )

    indicators: list[dict[str, Any]] = []
    for row in rows:
        scope, check = LOCAL_SCOPE.get(row["id"], ("na", ""))
        entry: dict[str, Any] = {
            "id": row["id"],
            "level": row["level"],
            "category": row["category"],
            "scope": scope,
        }
        if check:
            entry["check"] = check
        if remedy := REMEDIES.get(row["id"]):
            entry["remedy"] = {"do": remedy[0], "why": remedy[1]}
        entry["granularity"] = row["granularity"]
        if row["rda_ref"]:
            entry["rda_ref"] = row["rda_ref"]
        if row["fairsfair_ref"]:
            entry["fairsfair_ref"] = row["fairsfair_ref"]
        entry["text"] = row["text"]
        indicators.append(entry)

    return {
        "source": SOURCE,
        "license": SOURCE["license"],
        "levels": dict(_LEVELS),
        "scoring": _load_scoring(known),
        "indicators": indicators,
    }


def format_report(data: dict[str, Any]) -> str:
    """A per-level summary of what the model holds and how much of it we assess."""
    indicators = data["indicators"]
    levels = data["levels"]
    scope_of = {i["id"]: i["scope"] for i in indicators}

    lines: list[str] = []
    lines.append("FAIRplus Dataset Maturity Model (DSM) v1.2 — indicator coverage")
    lines.append("=" * 72)
    lines.append(
        f"{'Level':<5} {'Name':<26} {'C':>4} {'R':>4} {'H':>4} "
        f"{'total':>6} {'assessed':>9} {'na':>4}"
    )
    lines.append("-" * 72)

    for lvl in sorted(levels):
        at = [i for i in indicators if i["level"] == lvl]
        if not at:
            continue
        cats = {c: sum(1 for i in at if i["category"] == c) for c in "CRH"}
        assessed = sum(1 for i in at if scope_of[i["id"]] in ("full", "partial"))
        lines.append(
            f"{lvl:<5} {levels[lvl][:26]:<26} {cats['C']:>4} {cats['R']:>4} "
            f"{cats['H']:>4} {len(at):>6} {assessed:>9} {len(at) - assessed:>4}"
        )

    total = len(indicators)
    full = sum(1 for i in indicators if i["scope"] == "full")
    partial = sum(1 for i in indicators if i["scope"] == "partial")
    na = total - full - partial
    lines.append("-" * 72)
    lines.append(
        f"{'ALL':<5} {'':<26} "
        f"{sum(1 for i in indicators if i['category'] == 'C'):>4} "
        f"{sum(1 for i in indicators if i['category'] == 'R'):>4} "
        f"{sum(1 for i in indicators if i['category'] == 'H'):>4} "
        f"{total:>6} {full + partial:>9} {na:>4}"
    )
    lines.append("")
    lines.append(f"  full     {full:>3}  directly decidable from the assembled @graph")
    lines.append(f"  partial  {partial:>3}  presence/structure heuristic")
    lines.append(f"  na       {na:>3}  not assessable from a crate — reported, never failed")

    # A DISJOINT partition of the `na` set — the three groups overlap (Level 0 and
    # Level 5 each contain hosting indicators), so counting them independently would
    # sum past the total and misstate the coverage.
    ladder = [i for i in indicators if i["level"] not in (0, 5)]
    level0 = sum(1 for i in indicators if i["level"] == 0)
    level5 = sum(1 for i in indicators if i["level"] == 5)
    hosting_on_ladder = sum(1 for i in ladder if i["category"] == "H")
    other = sum(
        1 for i in ladder if i["category"] != "H" and i["scope"] == "na"
    )
    linked = sum(1 for i in indicators if i.get("rda_ref"))
    lines.append("")
    lines.append(f"  the {na} na, partitioned:")
    lines.append(f"    {level0:>3}  Level 0 — pre-FAIRification states, not a target")
    lines.append(f"    {level5:>3}  Level 5 — enterprise data governance, no crate can evidence it")
    lines.append(f"    {hosting_on_ladder:>3}  hosting-environment indicators at levels 1-4")
    lines.append(f"    {other:>3}  content/representation at levels 1-4 with no check yet")
    lines.append("")
    lines.append(f"  {linked} indicators carry the model's own RDA cross-reference.")
    lines.append(
        f"  Ceiling for any crate: level {max(lvl for lvl in levels if lvl not in (0, 5))}"
        " — levels 0 and 5 are off the ladder."
    )
    return "\n".join(lines)


def main() -> None:
    data = build_data()
    body = yaml.safe_dump(data, sort_keys=False, allow_unicode=True, width=100)
    OUT.write_text(_HEADER + "\n" + body)
    print(format_report(data))
    print()
    print(f"Wrote {OUT.relative_to(REPO)} ({len(data['indicators'])} indicators).")


if __name__ == "__main__":
    main()
