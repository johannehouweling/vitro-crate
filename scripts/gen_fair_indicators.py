#!/usr/bin/env python3
"""Generate ``fair/indicators.yaml`` from the vendored RDA FAIR Data Maturity Model.

Issue #356. The generic FAIR indicator *definitions* (dimension, priority, text) are
sourced from the canonical RDA FAIR Data Maturity Model spreadsheet
(``fair/rda_fdmm.xlsx``, DOI 10.15497/rda00050, CC-BY-4.0) rather than hand-authored,
so they cannot silently drift. The *local* decision — which indicators this tool can
assess intrinsically from a single RO-Crate, and with which check function — lives in
:data:`LOCAL_SCOPE` below; that is inherently repo-specific and stays here.

The workbook's own **scoring** is now read too, into a ``scoring:`` block — not to
compute it, but so that what this tool substitutes is stated rather than implied.
That third-party FAIR evaluators (F-UJI, the FAIR Evaluator, FAIROS) need a published,
resolvable URL and cannot score a pre-publication crate is true and irrelevant: the
instrument vendored two directories away needs neither, and :func:`_load_scoring`
records what it computes and why this tool does not.

The FAIRplus DSM ladder (``fair/dsm_indicators.yaml``) is generated from its own
vendored assessment workbook by ``scripts/gen_dsm_indicators.py`` — not touched here.

**All 41 published indicators are carried**, including the ones this tool cannot
answer; each of those states why. Carrying only the assessable subset, as this
generator used to, let a coverage figure overstate itself — "14 of 21 met" reads as
coverage of the model when it is really 14 of 41 with 20 never asked. Which parts of
a published instrument a tool answers is a local decision, and it has to be visible
rather than implied by absence. :func:`build_data` refuses to emit unless every
published indicator has either a check or a stated reason.

Regenerate with::

    uv run python scripts/gen_fair_indicators.py

``tests/test_fair_indicators_source.py`` asserts the committed file equals this
generator's output and that every indicator matches its RDA row.
"""

from __future__ import annotations

import pathlib
from typing import Any

import openpyxl
import yaml

REPO = pathlib.Path(__file__).resolve().parents[1]
RDA_XLSX = REPO / "fair" / "rda_fdmm.xlsx"
OUT = REPO / "fair" / "indicators.yaml"

# RDA indicator sheet: header row 0; columns PRINCIPLE(3), INDICATOR_ID(4),
# INDICATORS/text(5), PRIORITY(6).
_SHEET = "FAIR Indicators_v0.05"
# The workbook's other three sheets, all of which the scoring model spans.
_CALC_SHEET = "calc"
_LEVELS_SHEET = "LEVELS"
_DATA_SHEET = "DATA"

# Every published indicator is carried, in report order: indicator id ->
# (check-function name, reason-it-is-not-checked). Exactly one of the pair is set.
#
# An indicator with a check is scored; one with a reason is reported `out_of_scope` —
# never failed, and never silently dropped. Omitting the ones we cannot answer, as
# this file used to, made "14 of 21 met" read as coverage of the model when it was
# really 14 of 41 with 20 never asked. Which parts a tool answers is a local decision
# and has to be visible; the DSM and Bridge2AI generators carry their whole
# instruments for the same reason.
_PROTOCOL = (
    "the access protocol is a property of the repository serving the crate, not of "
    "the crate itself"
)
_HOSTING = "harvesting and indexing are properties of the repository serving the crate"
_PAYLOAD_PID = "the crate's payload files carry no persistent identifier of their own"
_PAYLOAD_SEMANTICS = (
    "the payload is tabular measurement data, not a knowledge representation; only "
    "the metadata layer is expressed in a standardised format"
)
_PAYLOAD_LINKS = "references made from inside the payload files are not inspected"

LOCAL_SCOPE: dict[str, tuple[str | None, str | None]] = {
    # ---- Findable ----------------------------------------------------------
    "RDA-F1-01M": ("pid_form", None),
    "RDA-F1-01D": (None, _PAYLOAD_PID),
    "RDA-F1-02M": ("root_global_id", None),
    "RDA-F1-02D": ("every_entity_has_id", None),
    "RDA-F2-01M": ("rich_metadata", None),
    "RDA-F3-01M": ("metadata_refs_data", None),
    "RDA-F4-01M": (None, _HOSTING),
    # ---- Accessible — every one is protocol / repository level --------------
    "RDA-A1-01M": (None, _PROTOCOL),
    "RDA-A1-02M": (None, _PROTOCOL),
    "RDA-A1-02D": (None, _PROTOCOL),
    "RDA-A1-03M": (None, _PROTOCOL),
    "RDA-A1-03D": (None, _PROTOCOL),
    "RDA-A1-04M": (None, _PROTOCOL),
    "RDA-A1-04D": (None, _PROTOCOL),
    "RDA-A1-05D": (None, _PROTOCOL),
    "RDA-A1.1-01M": (None, _PROTOCOL),
    "RDA-A1.1-01D": (None, _PROTOCOL),
    "RDA-A1.2-01D": (None, _PROTOCOL),
    "RDA-A2-01M": (None, _PROTOCOL),
    # ---- Interoperable -----------------------------------------------------
    "RDA-I1-01M": ("jsonld_context", None),
    "RDA-I1-02M": ("jsonld_context", None),
    "RDA-I1-01D": (None, _PAYLOAD_SEMANTICS),
    "RDA-I1-02D": (None, _PAYLOAD_SEMANTICS),
    "RDA-I2-01M": ("fair_vocabularies", None),
    "RDA-I2-01D": (None, _PAYLOAD_SEMANTICS),
    "RDA-I3-01M": ("qualified_refs", None),
    "RDA-I3-01D": (None, _PAYLOAD_LINKS),
    "RDA-I3-02M": (None, _PAYLOAD_LINKS),
    "RDA-I3-02D": (None, _PAYLOAD_LINKS),
    "RDA-I3-03M": ("qualified_refs", None),
    "RDA-I3-04M": (None, _PAYLOAD_LINKS),
    # ---- Reusable ----------------------------------------------------------
    "RDA-R1-01M": ("reuse_attributes", None),
    "RDA-R1.1-01M": ("license_present", None),
    "RDA-R1.1-02M": ("license_standard", None),
    "RDA-R1.1-03M": ("license_machine", None),
    "RDA-R1.2-01M": ("provenance", None),
    "RDA-R1.2-02M": (
        None,
        "no cross-community provenance language (PROV-O, PAV) is emitted, so there "
        "is nothing to check compliance against",
    ),
    "RDA-R1.3-01M": ("conforms_to_profile", None),
    "RDA-R1.3-02M": ("conforms_to_profile", None),
    # R1.3-01D delegates to MIT in-vitro coverage (the community reporting
    # standard); see builder/tools/mit_assessment.py and issue #313.
    "RDA-R1.3-01D": ("mit_coverage", None),
    "RDA-R1.3-02D": (
        None,
        "the payload's compliance with a machine-understandable community standard "
        "is a property of the data files, which this tool does not validate here",
    ),
}

SOURCES: dict[str, Any] = {
    # The specification and the journal article are DIFFERENT works and are recorded
    # separately, as the DSM block does. They were previously one entry — a `citation`
    # naming Bahim et al. beside a `doi` pointing at the RDA deliverable, which has no
    # authors — so anyone citing from here produced the right author list against the
    # wrong DOI.
    "rda": {
        "name": "RDA FAIR Data Maturity Model",
        "specification": {
            "name": "FAIR Data Maturity Model: specification and guidelines",
            "publisher": "Research Data Alliance",
            "doi": "10.15497/rda00050",
            "url": "https://doi.org/10.15497/rda00050",
            "year": 2020,
        },
        "peer_reviewed": {
            "citation": (
                "Bahim C, Casorrán-Amilburu C, Dekkers M, Herczog E, Loozen N, "
                "Repanas K, Russell K, Stall S. The FAIR Data Maturity Model: An "
                "Approach to Harmonise FAIR Assessments. Data Science Journal "
                "19(1):41 (2020)"
            ),
            "doi": "10.5334/dsj-2020-041",
        },
        "distribution": (
            "FAIR_evaluation_levels_v0.02.xlsx (sheet 'FAIR Indicators_v0.05'), "
            "Zenodo record 3909563 — vendored as fair/rda_fdmm.xlsx, byte-identical "
            "to the published file (md5 b6231346dff874a3675747ab3e295fd9)"
        ),
        "license": "CC-BY-4.0",
        "retrieved": "2026-07-25",
    },
    "dsm": {
        "name": "FAIRplus Dataset Maturity Model (DSM)",
        "url": "https://fairplus.github.io/Data-Maturity/",
        "note": (
            "Generated from the vendored assessment workbook by "
            "scripts/gen_dsm_indicators.py into fair/dsm_indicators.yaml, and scored "
            "by builder/tools/fair_assessment.py."
        ),
    },
    "nsdra": {
        "name": "Nanosafety data reusability (NSDRA)",
        "citation": "Ammar, Evelo & Willighagen 2024, Scientific Data 11:503",
        "doi": "10.1038/s41597-024-03324-x",
    },
}

_HEADER = """\
# GENERATED FILE — do not edit by hand.
#
# Regenerate with:  uv run python scripts/gen_fair_indicators.py
#
# The FAIR indicator definitions (dimension / priority / text) are derived from the
# vendored RDA FAIR Data Maturity Model spreadsheet (fair/rda_fdmm.xlsx,
# doi:10.15497/rda00050, CC-BY-4.0). The local scope/check mapping (which indicators
# are assessed intrinsically from one RO-Crate, and with which check) lives in
# scripts/gen_fair_indicators.py. Accessibility indicators are protocol/repository
# level and reported out-of-scope, not failed. R1.3-01D delegates to MIT
# in-vitro coverage. The FAIRplus DSM ladder is generated separately into
# fair/dsm_indicators.yaml by scripts/gen_dsm_indicators.py.
#
# ALL 41 published indicators are carried, including the ones this tool cannot answer.
# Each of those states WHY, as `scope: out_of_scope` plus a `reason`. Omitting them
# made a coverage figure overstate itself: "14 of 21 met" reads as coverage of the
# model when it is really 14 of 41 with 20 never asked.
"""


def _load_rda() -> dict[str, dict[str, str]]:
    """Canonical RDA indicator definitions keyed by indicator id."""
    wb = openpyxl.load_workbook(RDA_XLSX, read_only=True, data_only=True)
    ws = wb[_SHEET]
    out: dict[str, dict[str, str]] = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        iid = row[4]
        if iid and str(iid).startswith("RDA-"):
            out[str(iid)] = {
                "dimension": str(row[3])[0],
                "priority": str(row[6]).strip().lower(),
                "text": str(row[5]).strip(),
            }
    return out


def _load_scoring() -> dict[str, Any]:
    """What the workbook computes from the 41 answers — read from its cells, not restated.

    Two things, and this tool reproduces exactly one of them.

    **The per-indicator boolean is the workbook's own.** Its native answer is a five-way
    maturity metric (column H, a dropdown over ``DATA!L1:L5``), but column J collapses it
    all-or-nothing: ``1`` only for "4 - fully implemented", ``0`` for every partial state
    *and* for a blank. That is precisely what a met/failed verdict here means, so the
    boolean is not a coarsening of the instrument — it is the instrument's own column.
    The 0-4 shading survives only in the workbook's radar pictures.

    **The ladder is not reproduced.** ``calc!C13:F13`` computes a maturity level per FAIR
    area from the met/total counts per priority tier, and this tool publishes no level at
    all. Recording the formula verbatim rather than a parsed threshold table is
    deliberate: one fact stated twice is how a generated file and its source drift.
    """
    wb = openpyxl.load_workbook(RDA_XLSX, data_only=False)
    ind, calc = wb[_SHEET], wb[_CALC_SHEET]
    levels_sheet, data = wb[_LEVELS_SHEET], wb[_DATA_SHEET]

    metric = [str(data.cell(row=r, column=12).value) for r in range(1, 6)]
    priorities = [str(data.cell(row=r, column=3).value) for r in range(1, 4)]
    areas = [str(calc.cell(row=5, column=c).value) for c in range(3, 7)]
    levels = [
        {
            "name": str(data.cell(row=r, column=5).value),
            "definition": str(levels_sheet.cell(row=r + 4, column=5).value),
        }
        for r in range(1, 7)
    ]
    per_indicator, ladder = str(ind["J2"].value), str(calc["C13"].value)

    # Generation fails loudly rather than emitting a model the sheet does not hold.
    if not all(f"DATA!$L${n}" in per_indicator for n in (1, 2, 3, 4)):
        raise ValueError(f"J2 no longer collapses the four sub-4 metrics: {per_indicator}")
    if not all(f"DATA!$E${n}" in ladder for n in range(1, 7)):
        raise ValueError(f"calc!C13 no longer names all six levels: {ladder}")
    if [lvl["name"] for lvl in levels] != [f"Level {n}" for n in range(6)]:
        raise ValueError(f"DATA!E1:E6 is not the six levels: {levels}")
    if len(metric) != 5 or len(priorities) != 3 or len(areas) != 4:
        raise ValueError("the sheet's metric / priority / area vocabularies changed")

    return {
        "metric": metric,
        "per_indicator": {
            "cell": f"'{_SHEET}'!J2",
            "formula": per_indicator,
            "note": (
                "Only '4 - fully implemented' scores 1. Every partial state scores 0, and "
                "so does a blank — the trailing $H2=0 term. This tool's met/failed verdict "
                "is this column, not a coarsening of it."
            ),
        },
        "ladder": {
            "cell": f"{_CALC_SHEET}!C13:F13",
            "formula": ladder,
            "areas": areas,
            "priorities": priorities,
            "levels": levels,
        },
        "published_output": (
            "one maturity level per FAIR area, from the met and total counts per priority "
            "tier within that area (calc!C6:F11)"
        ),
        "local_output": "a met / failed / not-assessed count over all 41 indicators, and no level",
        "deviation_note": (
            "The ladder gates every level above 0 on ALL essential indicators in the area "
            "being met, and an indicator this tool reports out_of_scope can never be met. "
            "Measured on the model as carried: Accessibility is 12 of 12 out_of_scope, so "
            "its level would read 0 on every crate forever; Findability carries two "
            "out_of_scope essentials (RDA-F1-01D, RDA-F4-01M), so it could not reach Level "
            "1 on a perfect crate. Publishing the ladder would report the hosting "
            "repository's properties as the crate's failure, which is the reading this "
            "tool refuses everywhere else. The substitution is a flat count — declared "
            "here rather than left to be inferred from a missing key."
        ),
    }


def build_data() -> dict[str, Any]:
    """The full ``indicators.yaml`` payload: sources + all 41 RDA indicators."""
    rda = _load_rda()
    unknown = sorted(set(LOCAL_SCOPE) - set(rda))
    if unknown:
        raise KeyError(
            f"{unknown} is not an RDA FAIR Data Maturity Model indicator. "
            "Fix the mapping rather than the model."
        )
    missing = sorted(set(rda) - set(LOCAL_SCOPE))
    if missing:
        raise KeyError(
            f"the published model defines indicators this generator does not place: "
            f"{missing}. Every one needs either a check or a stated reason — silence "
            "is how a coverage figure comes to overstate itself."
        )

    indicators: list[dict[str, Any]] = []
    for iid, (check, reason) in LOCAL_SCOPE.items():
        if bool(check) == bool(reason):
            raise ValueError(
                f"{iid}: set exactly one of check / reason — an indicator is either "
                "scored or explained, never both and never neither."
            )
        d = rda[iid]
        entry: dict[str, Any] = {
            "id": iid,
            "dimension": d["dimension"],
            "priority": d["priority"],
        }
        if check:
            entry["check"] = check
        else:
            entry["scope"] = "out_of_scope"
            entry["reason"] = reason
        entry["text"] = d["text"]
        indicators.append(entry)
    return {"sources": SOURCES, "scoring": _load_scoring(), "indicators": indicators}


def format_report(data: dict[str, Any]) -> str:
    """What the model asks, and how much of it this tool answers."""
    inds = data["indicators"]
    scored = [i for i in inds if "check" in i]
    lines = [
        "RDA FAIR Data Maturity Model — indicator coverage",
        "=" * 66,
        f"{'Dim':<5} {'total':>6} {'scored':>7} {'out of scope':>13}",
        "-" * 66,
    ]
    for dim in sorted({i["dimension"] for i in inds}):
        at = [i for i in inds if i["dimension"] == dim]
        n = sum(1 for i in at if "check" in i)
        lines.append(f"{dim:<5} {len(at):>6} {n:>7} {len(at) - n:>13}")
    lines += [
        "-" * 66,
        f"{'ALL':<5} {len(inds):>6} {len(scored):>7} {len(inds) - len(scored):>13}",
        "",
        "  by priority:",
    ]
    for pri in ("essential", "important", "useful"):
        at = [i for i in inds if i["priority"] == pri]
        n = sum(1 for i in at if "check" in i)
        lines.append(f"    {pri:<10} {n:>2} scored of {len(at):>2}")
    lines += ["", "  why the rest are out of scope:"]
    reasons: dict[str, list[str]] = {}
    for i in inds:
        if r := i.get("reason"):
            reasons.setdefault(r, []).append(i["id"])
    for reason, ids in sorted(reasons.items(), key=lambda kv: -len(kv[1])):
        lines.append(f"    {len(ids):>2}  {reason}")
        lines.append(f"        {', '.join(ids)}")
    return "\n".join(lines)


def main() -> None:
    data = build_data()
    body = yaml.safe_dump(data, sort_keys=False, allow_unicode=True, width=100)
    OUT.write_text(_HEADER + "\n" + body)
    print(format_report(data))
    print()
    print(f"Wrote {OUT.relative_to(REPO)} ({len(LOCAL_SCOPE)} indicators from RDA).")


if __name__ == "__main__":
    main()
